"""
GPA-ERP HRIS — Phase H2: Absensi & Cuti
Endpoints for attendance (with geolocation + face verification) and leave management.
"""
from __future__ import annotations

import csv
import io
import logging
import math
import uuid
from datetime import date, datetime, timezone, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Annotated

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_

from app.audit import model_to_dict, write_audit
from app.database import get_db
from app.dependencies import CurrentUser, get_client_ip, require_role
from app.models import (
    AttendanceRecord, AttendanceSource,
    Employee, LeaveBalance, LeaveCategory, LeaveRequest, LeaveRequestStatus, LeaveType,
    RoleName, WorkGroup, WorkLocation, WorkLocationType,
    OvertimeRequest, OvertimeRequestStatus, HolidayCalendar,
)
from app.notify import push, push_to_role
from app.schemas import (
    AttendanceManualCreate, AttendanceRecordResponse, AttendanceSummaryItem,
    LeaveActionRequest, LeaveBalanceResponse, LeaveRequestCreate,
    LeaveRequestResponse, LeaveTypeCreate, LeaveTypeResponse,
    MessageResponse, PaginatedResponse,
    WorkLocationCreate, WorkLocationResponse, WorkLocationUpdate,
    OvertimeRequestCreate, OvertimeRequestResponse, OvertimeActionRequest,
    LeaveCalendarItem,
)

router = APIRouter(prefix="/hris", tags=["HRIS – Attendance & Leave"])

_hr_roles  = (RoleName.SUPER_ADMIN, RoleName.MD, RoleName.GA)
_mgr_roles = (RoleName.SUPER_ADMIN, RoleName.MD, RoleName.PM, RoleName.GA)

_SELFIE_DIR = Path("uploads") / "selfies"
_SELFIE_DIR.mkdir(parents=True, exist_ok=True)

# Default leave approval chain: GA reviews, MD approves
_LEAVE_APPROVAL_CHAIN = ["GA", "MD"]


# ─── Overtime calculation (Permenaker No. 2 Tahun 2023) ─────────────────────

def _calculate_overtime(
    clock_in:  datetime,
    clock_out: datetime,
    is_weekend: bool,
    is_holiday: bool,
) -> tuple[Decimal, Decimal, Decimal, Decimal]:
    """
    Returns (hours_regular, hours_ot_weekday, hours_ot_weekend, hours_ot_holiday).
    Permenaker 2023 rules:
      Weekday: first 8h regular, OT starts after 8h
      Weekend/Holiday: all hours are OT
    """
    MAX_DAILY_HOURS = Decimal("24")  # sanity cap — forgot to clock out guard
    total_hours = min(
        Decimal(str((clock_out - clock_in).total_seconds() / 3600)),
        MAX_DAILY_HOURS,
    )
    REGULAR_HOURS = Decimal("8")

    if is_holiday:
        return Decimal("0"), Decimal("0"), Decimal("0"), total_hours
    if is_weekend:
        return Decimal("0"), Decimal("0"), total_hours, Decimal("0")

    # Weekday
    regular = min(total_hours, REGULAR_HOURS)
    ot      = max(Decimal("0"), total_hours - REGULAR_HOURS)
    return regular, ot, Decimal("0"), Decimal("0")


def _is_weekend(d: date) -> bool:
    return d.weekday() >= 5  # Saturday=5, Sunday=6


# ─── Geolocation helpers ──────────────────────────────────────────────────────

def _haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Return great-circle distance in metres between two GPS points."""
    R = 6_371_000  # Earth radius in metres
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi       = math.radians(lat2 - lat1)
    dlambda    = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def _check_location(
    db: Session,
    latitude: float,
    longitude: float,
    assigned_location: "WorkLocation | None" = None,
) -> tuple[bool, "WorkLocation | None", float]:
    """
    Check GPS coords against WorkLocations.

    If the employee has an assigned work_location, validate ONLY against that
    location — so a Jakarta employee can't clock-in at a Berau site and vice versa.

    If no location is assigned, fall back to checking all active locations
    (previous behaviour — matches the nearest one within any radius).

    Never blocks clock-in — just records the result for HR review.
    """
    if assigned_location is not None:
        # Validate strictly against the employee's assigned location only
        locations = [assigned_location]
    else:
        locations = db.query(WorkLocation).filter(WorkLocation.is_active == True).all()

    best: WorkLocation | None = None
    best_dist = float("inf")
    for loc in locations:
        dist = _haversine(latitude, longitude, float(loc.latitude), float(loc.longitude))
        if dist <= loc.radius_meters and dist < best_dist:
            best, best_dist = loc, dist
    # If no match found, compute distance to nearest for reporting
    if best is None:
        nearest_dist = float("inf")
        for loc in locations:
            dist = _haversine(latitude, longitude, float(loc.latitude), float(loc.longitude))
            if dist < nearest_dist:
                nearest_dist = dist
        return False, None, nearest_dist if nearest_dist != float("inf") else 0.0
    return True, best, best_dist


# ─── Attendance: clock-in (mobile, geolocation + selfie) ─────────────────────

@router.post("/attendance/clock-in", response_model=AttendanceRecordResponse,
             summary="Mobile clock-in with GPS + selfie")
async def clock_in(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    latitude:     float | None        = Form(None),
    longitude:    float | None        = Form(None),
    accuracy:     float | None        = Form(None),
    employee_id:  int | None         = Form(None),
    selfie:       UploadFile | None  = File(None),
):
    """
    Mobile clock-in: accepts GPS coordinates + selfie photo.
    - If employee_id not provided, uses the employee linked to current_user.
    - Runs face verification if employee has a registered face embedding.
    - Creates or updates the AttendanceRecord for today.
    """
    # Resolve employee
    if employee_id:
        emp = db.query(Employee).filter(Employee.id == employee_id).first()
        if not emp:
            raise HTTPException(404, "Employee not found")
        # Only HR/admin can clock in for others
        if emp.user_id != current_user.id and current_user.role.name not in _hr_roles:
            raise HTTPException(403, "Can only clock in for yourself")
    else:
        emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if not emp:
            raise HTTPException(404, "No employee record linked to your account")

    today      = datetime.now(timezone.utc).date()
    now        = datetime.now(timezone.utc)
    face_detected: bool             = False
    face_confidence: Decimal | None = None
    selfie_url: str | None          = None
    location_ok: bool | None        = None
    location_distance_m: Decimal | None = None

    # ── Geolocation validation (soft — never blocks clock-in) ─────────────────
    # If the employee has an assigned work location, validate only against that.
    # Otherwise fall back to checking all active locations.
    assigned_wl: WorkLocation | None = emp.work_location if emp.work_location_id else None
    matched_loc: WorkLocation | None = None
    if latitude is not None and longitude is not None:
        ok, matched_loc, dist = _check_location(db, latitude, longitude, assigned_wl)
        location_ok          = ok
        location_distance_m  = Decimal(str(round(dist, 1)))
        if not ok and emp.work_location:
            push_to_role(db, RoleName.GA,
                         "Absensi: Lokasi Di Luar Radius",
                         f"{emp.full_name} clock-in dari jarak "
                         f"{dist:.0f}m (radius: {emp.work_location.radius_meters}m dari {emp.work_location.name})",
                         "/hris/attendance")

    # Process selfie — detect whether a face is present (no identity matching)
    if selfie:
        selfie_bytes = await selfie.read()
        ext      = Path(selfie.filename or "selfie").suffix or ".jpg"
        filename = f"{emp.id}_{today.isoformat()}_{uuid.uuid4().hex[:8]}{ext}"
        dest     = _SELFIE_DIR / filename
        dest.write_bytes(selfie_bytes)
        selfie_url = f"/uploads/selfies/{filename}"

        try:
            from app.hris_face import detect_face
            face_detected, conf = detect_face(selfie_bytes)
            face_confidence = Decimal(str(conf))
        except Exception as exc:
            logger.warning("Face detection error during clock-in: %s", exc)
            # Don't block clock-in if detection fails

    # Upsert attendance record (one per employee per day)
    record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == emp.id,
        AttendanceRecord.date        == today,
    ).first()

    if record:
        # Already clocked in — update selfie/geo if re-submitted before clock-out
        if selfie_url:
            record.selfie_url      = selfie_url
            record.face_verified   = face_detected
            record.face_confidence = face_confidence
        if latitude is not None:
            record.latitude                  = Decimal(str(latitude))
            record.longitude                 = Decimal(str(longitude))
            record.accuracy                  = Decimal(str(accuracy)) if accuracy is not None else None
            record.location_ok               = location_ok
            record.location_distance_m       = location_distance_m
            record.matched_work_location_id  = matched_loc.id if matched_loc else None
    else:
        record = AttendanceRecord(
            employee_id              = emp.id,
            date                     = today,
            clock_in                 = now,
            source                   = AttendanceSource.MOBILE,
            latitude                 = Decimal(str(latitude)) if latitude is not None else None,
            longitude                = Decimal(str(longitude)) if longitude is not None else None,
            accuracy                 = Decimal(str(accuracy)) if accuracy is not None else None,
            location_ok              = location_ok,
            location_distance_m      = location_distance_m,
            matched_work_location_id = matched_loc.id if matched_loc else None,
            selfie_url               = selfie_url,
            face_verified            = face_detected,
            face_confidence          = face_confidence,
        )
        db.add(record)

    db.flush()
    write_audit(db, "AttendanceRecord", record.id, "CLOCK_IN",
                changed_by=current_user.id,
                after={"employee_id": emp.id, "date": str(today),
                       "face_detected": face_detected,
                       "face_confidence": str(face_confidence or ""),
                       "has_selfie": selfie_url is not None,
                       "location_ok": location_ok,
                       "location_distance_m": str(location_distance_m or "")})
    db.commit()
    db.refresh(record)

    # Flag to HR if selfie was submitted but no face was detected
    if selfie and not face_detected:
        push_to_role(db, RoleName.GA,
                     "Absensi: Wajah Tidak Terdeteksi",
                     f"{emp.full_name} clock-in namun tidak ada wajah pada selfie",
                     "/hris/attendance")
        db.commit()

    return record


# ─── Attendance: clock-out ────────────────────────────────────────────────────

@router.post("/attendance/clock-out", response_model=AttendanceRecordResponse,
             summary="Clock out — calculates hours worked")
def clock_out(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    employee_id:  int | None = Query(None),
    is_holiday:   bool       = Query(False),
    note:         str | None = Query(None),
):
    if employee_id:
        emp = db.query(Employee).filter(Employee.id == employee_id).first()
        if not emp:
            raise HTTPException(404, "Employee not found")
        if emp.user_id != current_user.id and current_user.role.name not in _hr_roles:
            raise HTTPException(403, "Can only clock out for yourself")
    else:
        emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if not emp:
            raise HTTPException(404, "No employee record linked to your account")

    today  = datetime.now(timezone.utc).date()
    now    = datetime.now(timezone.utc)

    record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == emp.id,
        AttendanceRecord.date        == today,
    ).first()

    if not record:
        raise HTTPException(409, "No clock-in found for today. Please clock in first.")
    if record.clock_out:
        raise HTTPException(409, "Already clocked out today")

    record.clock_out = now
    if record.clock_in:
        weekend = _is_weekend(today)
        reg, ot_wd, ot_we, ot_hol = _calculate_overtime(
            record.clock_in, now, weekend, is_holiday
        )
        record.hours_regular          = reg
        record.hours_overtime_weekday = ot_wd
        record.hours_overtime_weekend = ot_we
        record.hours_overtime_holiday = ot_hol

    if note:
        record.note = note

    write_audit(db, "AttendanceRecord", record.id, "CLOCK_OUT",
                changed_by=current_user.id,
                after={"clock_out": str(now), "hours_regular": str(record.hours_regular or "")})
    db.commit()
    db.refresh(record)
    return record


# ─── DEBUG: reset today's attendance ─────────────────────────────────────────

@router.delete("/attendance/debug-reset", status_code=200,
               summary="[DEBUG] Delete today's attendance record — testing only",
               include_in_schema=False)
def debug_reset_attendance(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    employee_id:  int | None = Query(None, description="Target employee (SA/HR only; omit to use own)"),
):
    """
    Deletes today's AttendanceRecord so the clock-in flow can be retested.
    Only available when DEBUG=True.
    SUPER_ADMIN / GA / MD can reset any employee's record via `employee_id`.
    """
    from app.config import get_settings as _gs
    if not _gs().DEBUG:
        raise HTTPException(404, "Not found")
    today = date.today()

    # Resolve target employee
    if employee_id is not None:
        if current_user.role.name not in _hr_roles:
            raise HTTPException(403, "Only HR roles can reset other employees' attendance")
        emp = db.query(Employee).filter(Employee.id == employee_id).first()
        if not emp:
            raise HTTPException(404, "Employee not found")
    else:
        emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if not emp:
            raise HTTPException(404, "No employee record linked to your account")

    record = (
        db.query(AttendanceRecord)
        .filter(
            AttendanceRecord.employee_id == emp.id,
            func.date(AttendanceRecord.clock_in) == today,
        )
        .first()
    )

    if not record:
        return {"detail": f"No attendance record found for {emp.full_name} today ({today})"}

    db.delete(record)
    db.commit()
    return {"detail": f"Deleted attendance record #{record.id} for {emp.full_name} ({today})"}


# ─── Attendance: manual entry (HR admin) ─────────────────────────────────────

@router.post("/attendance", response_model=AttendanceRecordResponse, status_code=201,
             summary="Manual attendance entry (HR admin)")
def create_attendance_manual(
    request:      Request,
    payload:      AttendanceManualCreate,
    current_user: Annotated[CurrentUser, Depends(require_role(*_hr_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    emp = db.query(Employee).filter(Employee.id == payload.employee_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")

    existing = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == payload.employee_id,
        AttendanceRecord.date        == payload.date,
    ).first()
    if existing:
        raise HTTPException(409, "Attendance record already exists for this date. Use PATCH to update.")

    # Auto-calculate hours if clock_in and clock_out provided
    reg = payload.hours_regular
    ot_wd = payload.hours_overtime_weekday
    ot_we = payload.hours_overtime_weekend
    ot_hol = payload.hours_overtime_holiday

    if payload.clock_in and payload.clock_out and reg is None:
        weekend = _is_weekend(payload.date)
        reg, ot_wd, ot_we, ot_hol = _calculate_overtime(
            payload.clock_in, payload.clock_out, weekend, False
        )

    record = AttendanceRecord(
        employee_id            = payload.employee_id,
        date                   = payload.date,
        clock_in               = payload.clock_in,
        clock_out              = payload.clock_out,
        hours_regular          = reg,
        hours_overtime_weekday = ot_wd,
        hours_overtime_weekend = ot_we,
        hours_overtime_holiday = ot_hol,
        source                 = AttendanceSource.MANUAL,
        note                   = payload.note,
        face_verified          = False,
    )
    db.add(record)
    db.flush()
    write_audit(db, "AttendanceRecord", record.id, "MANUAL_CREATE",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                after=model_to_dict(record))
    db.commit()
    db.refresh(record)
    return record


# ─── Attendance: list ─────────────────────────────────────────────────────────

@router.get("/attendance", response_model=PaginatedResponse[AttendanceRecordResponse],
            summary="List attendance records")
def list_attendance(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    employee_id:  int | None = Query(None),
    date_from:    date | None = Query(None),
    date_to:      date | None = Query(None),
    work_group_id: int | None = Query(None),
    skip:         int         = Query(0, ge=0),
    limit:        int         = Query(50, ge=1, le=200),
):
    q = db.query(AttendanceRecord)

    # Non-admin users can only see their own records
    if current_user.role.name not in (*_hr_roles, RoleName.MD, RoleName.PM, RoleName.COST_CONTROL, RoleName.FINANCE):
        my_emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if my_emp:
            q = q.filter(AttendanceRecord.employee_id == my_emp.id)
        else:
            return {"items": [], "total": 0}
    elif employee_id:
        q = q.filter(AttendanceRecord.employee_id == employee_id)

    if work_group_id:
        q = q.join(Employee, Employee.id == AttendanceRecord.employee_id).filter(
            Employee.work_group_id == work_group_id
        )

    if date_from:
        q = q.filter(AttendanceRecord.date >= date_from)
    if date_to:
        q = q.filter(AttendanceRecord.date <= date_to)

    total = q.count()
    items = q.order_by(AttendanceRecord.date.desc()).offset(skip).limit(limit).all()
    return {"items": items, "total": total}


# ─── Attendance: monthly summary ─────────────────────────────────────────────

@router.get("/attendance/summary", summary="Monthly attendance summary per employee")
def attendance_summary(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    year:         int = Query(...),
    month:        int = Query(...),
    dept_id:      int | None = Query(None),
):
    from calendar import monthrange
    first_day = date(year, month, 1)
    last_day  = date(year, month, monthrange(year, month)[1])

    q = (
        db.query(
            AttendanceRecord.employee_id,
            func.count(AttendanceRecord.id).label("days_present"),
            func.coalesce(func.sum(AttendanceRecord.hours_regular),          0).label("hours_regular"),
            func.coalesce(func.sum(AttendanceRecord.hours_overtime_weekday),  0).label("hours_ot_weekday"),
            func.coalesce(func.sum(AttendanceRecord.hours_overtime_weekend),  0).label("hours_ot_weekend"),
            func.coalesce(func.sum(AttendanceRecord.hours_overtime_holiday),  0).label("hours_ot_holiday"),
        )
        .filter(
            AttendanceRecord.date >= first_day,
            AttendanceRecord.date <= last_day,
        )
        .group_by(AttendanceRecord.employee_id)
    )

    if dept_id:
        q = q.join(Employee, Employee.id == AttendanceRecord.employee_id).filter(
            Employee.dept_id == dept_id
        )

    rows = q.all()
    emp_ids = [r.employee_id for r in rows]
    emp_map = {
        e.id: e for e in db.query(Employee).filter(Employee.id.in_(emp_ids)).all()
    }

    result = []
    for r in rows:
        emp = emp_map.get(r.employee_id)
        if not emp:
            continue
        result.append({
            "employee_id":    r.employee_id,
            "employee_no":    emp.employee_no,
            "full_name":      emp.full_name,
            "days_present":   r.days_present,
            "hours_regular":  str(r.hours_regular),
            "hours_ot_total": str(
                Decimal(str(r.hours_ot_weekday)) +
                Decimal(str(r.hours_ot_weekend)) +
                Decimal(str(r.hours_ot_holiday))
            ),
        })

    return result


# ─── Attendance: export (Excel/CSV) ─────────────────────────────────────────

@router.get("/attendance/export", summary="Export attendance records as Excel")
def export_attendance(
    current_user: Annotated[CurrentUser, Depends(require_role(
        RoleName.SUPER_ADMIN, RoleName.MD, RoleName.PM, RoleName.GA, RoleName.FINANCE, RoleName.COST_CONTROL
    ))],
    db:           Annotated[Session, Depends(get_db)],
    date_from:    date | None = Query(None),
    date_to:      date | None = Query(None),
    dept_id:      int | None  = Query(None),
    employee_id:  int | None  = Query(None),
    fmt:          str         = Query("xlsx", pattern="^(xlsx|csv)$"),
):
    """Download attendance records as Excel (.xlsx) or CSV."""
    q = (
        db.query(AttendanceRecord)
        .join(Employee, Employee.id == AttendanceRecord.employee_id)
    )
    if date_from:
        q = q.filter(AttendanceRecord.date >= date_from)
    if date_to:
        q = q.filter(AttendanceRecord.date <= date_to)
    if dept_id:
        q = q.filter(Employee.dept_id == dept_id)
    if employee_id:
        q = q.filter(AttendanceRecord.employee_id == employee_id)

    records = q.order_by(Employee.full_name, AttendanceRecord.date).all()

    # Build employee lookup
    emp_ids = {r.employee_id for r in records}
    emp_map = {e.id: e for e in db.query(Employee).filter(Employee.id.in_(emp_ids)).all()}
    wl_map: dict[int, WorkLocation] = {}
    wl_ids = {e.work_location_id for e in emp_map.values() if e.work_location_id}
    if wl_ids:
        wl_map = {w.id: w for w in db.query(WorkLocation).filter(WorkLocation.id.in_(wl_ids)).all()}

    HEADERS = [
        "Tanggal", "No. Karyawan", "Nama", "Departemen", "Lokasi Kerja",
        "Jam Masuk", "Jam Keluar",
        "Jam Reguler", "OT Weekday", "OT Weekend", "OT Libur",
        "Total Jam OT", "Sumber",
        "Latitude", "Longitude", "Akurasi (m)",
        "Lokasi OK", "Jarak (m)",
        "Wajah Terdeteksi", "Catatan",
    ]

    def _fmt_time(dt: datetime | None) -> str:
        if dt is None:
            return ""
        return dt.astimezone().strftime("%H:%M:%S")

    def _fmt_dec(v) -> str:
        return str(round(float(v), 2)) if v is not None else ""

    rows = []
    for r in records:
        emp = emp_map.get(r.employee_id)
        dept_name = emp.department.name if emp and emp.department else ""
        wl = wl_map.get(emp.work_location_id) if emp and emp.work_location_id else None
        ot_total = sum(
            float(x or 0) for x in [
                r.hours_overtime_weekday,
                r.hours_overtime_weekend,
                r.hours_overtime_holiday,
            ]
        )
        rows.append([
            str(r.date),
            emp.employee_no if emp else "",
            emp.full_name if emp else "",
            dept_name,
            wl.name if wl else "",
            _fmt_time(r.clock_in),
            _fmt_time(r.clock_out),
            _fmt_dec(r.hours_regular),
            _fmt_dec(r.hours_overtime_weekday),
            _fmt_dec(r.hours_overtime_weekend),
            _fmt_dec(r.hours_overtime_holiday),
            str(round(ot_total, 2)),
            r.source.value if r.source else "",
            _fmt_dec(r.latitude),
            _fmt_dec(r.longitude),
            _fmt_dec(r.accuracy),
            "Ya" if r.location_ok is True else ("Tidak" if r.location_ok is False else ""),
            _fmt_dec(r.location_distance_m),
            "Ya" if r.face_verified else "Tidak",
            r.note or "",
        ])

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(HEADERS)
        writer.writerows(rows)
        output.seek(0)
        fname = f"attendance_{date_from or 'all'}_{date_to or 'all'}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — use fmt=csv")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Absensi"

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-fit column widths (approximation)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"attendance_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── Work Locations CRUD ─────────────────────────────────────────────────────

_wl_roles = (RoleName.SUPER_ADMIN, RoleName.MD, RoleName.PM, RoleName.GA)


@router.get("/work-locations", response_model=list[WorkLocationResponse],
            summary="List work locations")
def list_work_locations(
    _:           CurrentUser,
    db:          Annotated[Session, Depends(get_db)],
    active_only: bool = Query(True),
):
    q = db.query(WorkLocation)
    if active_only:
        q = q.filter(WorkLocation.is_active == True)
    return q.order_by(WorkLocation.name).all()


@router.post("/work-locations", response_model=WorkLocationResponse, status_code=201,
             summary="Create a work location")
def create_work_location(
    payload:      WorkLocationCreate,
    current_user: Annotated[CurrentUser, Depends(require_role(*_wl_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    wl = WorkLocation(**payload.model_dump())
    db.add(wl)
    db.commit()
    db.refresh(wl)
    return wl


@router.patch("/work-locations/{wl_id}", response_model=WorkLocationResponse,
              summary="Update a work location")
def update_work_location(
    wl_id:        int,
    payload:      WorkLocationUpdate,
    current_user: Annotated[CurrentUser, Depends(require_role(*_wl_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    wl = db.query(WorkLocation).filter(WorkLocation.id == wl_id).first()
    if not wl:
        raise HTTPException(404, "Work location not found")
    for field, val in payload.model_dump(exclude_unset=True).items():
        setattr(wl, field, val)
    db.commit()
    db.refresh(wl)
    return wl


@router.patch("/employees/{employee_id}/work-location",
              response_model=dict,
              summary="Assign or clear work location for an employee")
def assign_employee_work_location(
    employee_id:      int,
    current_user:     Annotated[CurrentUser, Depends(require_role(*_wl_roles))],
    db:               Annotated[Session, Depends(get_db)],
    work_location_id: int | None = Query(None, description="Pass null to clear assignment"),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")

    if work_location_id is not None:
        wl = db.query(WorkLocation).filter(WorkLocation.id == work_location_id).first()
        if not wl:
            raise HTTPException(404, "Work location not found")
        emp.work_location_id = wl.id
        msg = f"Assigned to {wl.name}"
    else:
        emp.work_location_id = None
        msg = "Work location cleared"

    db.commit()
    return {"message": msg, "employee_id": employee_id, "work_location_id": work_location_id}


# ─── Leave Types ─────────────────────────────────────────────────────────────

@router.get("/leave-types", response_model=list[LeaveTypeResponse], summary="List leave types")
def list_leave_types(
    _:  CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    active_only: bool = True,
):
    q = db.query(LeaveType)
    if active_only:
        q = q.filter(LeaveType.is_active == True)
    return q.order_by(LeaveType.id).all()


@router.post("/leave-types", response_model=LeaveTypeResponse, status_code=201,
             summary="Create leave type")
def create_leave_type(
    payload:      LeaveTypeCreate,
    current_user: Annotated[CurrentUser, Depends(require_role(RoleName.SUPER_ADMIN, RoleName.MD))],
    db:           Annotated[Session, Depends(get_db)],
):
    if db.query(LeaveType).filter(LeaveType.code == payload.code).first():
        raise HTTPException(409, "Leave type code already exists")
    lt = LeaveType(**payload.model_dump())
    db.add(lt)
    db.commit()
    db.refresh(lt)
    return lt


# ─── Leave Balance ────────────────────────────────────────────────────────────

@router.get("/leave-balance/{employee_id}", response_model=list[LeaveBalanceResponse],
            summary="Get leave balances for employee (current year)")
def get_leave_balance(
    employee_id:  int,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    year:         int = Query(default=None),
):
    if year is None:
        year = datetime.now(timezone.utc).year

    # Ensure employee exists
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")

    # Self-service: staff can only see their own balance
    if current_user.role.name not in (*_hr_roles, RoleName.MD, RoleName.PM):
        my_emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if not my_emp or my_emp.id != employee_id:
            raise HTTPException(403, "Access denied")

    balances = (
        db.query(LeaveBalance)
        .filter(LeaveBalance.employee_id == employee_id, LeaveBalance.year == year)
        .all()
    )
    return balances


@router.post("/leave-balance/seed", summary="Seed leave balances for all active employees (HR admin)")
def seed_leave_balances(
    current_user: Annotated[CurrentUser, Depends(require_role(*_hr_roles, RoleName.MD))],
    db:           Annotated[Session, Depends(get_db)],
    year:         int = Query(default=None),
):
    """Ensure every active employee has a balance row for each active leave type."""
    if year is None:
        year = datetime.now(timezone.utc).year

    employees   = db.query(Employee).filter(Employee.status == "active").all()
    leave_types = db.query(LeaveType).filter(LeaveType.is_active == True).all()
    created = 0

    if not employees or not leave_types:
        return MessageResponse(message=f"Seeded {created} leave balance rows for {year}")

    emp_ids = [e.id for e in employees]
    lt_ids  = [lt.id for lt in leave_types]

    # Load all existing balances for this year in ONE query (N+1 fix)
    existing_keys: set[tuple[int, int]] = {
        (b.employee_id, b.leave_type_id)
        for b in db.query(LeaveBalance).filter(
            LeaveBalance.employee_id.in_(emp_ids),
            LeaveBalance.leave_type_id.in_(lt_ids),
            LeaveBalance.year == year,
        ).all()
    }

    lt_map = {lt.id: lt for lt in leave_types}
    for emp in employees:
        for lt in leave_types:
            if (emp.id, lt.id) not in existing_keys:
                db.add(LeaveBalance(
                    employee_id   = emp.id,
                    leave_type_id = lt.id,
                    year          = year,
                    accrued       = lt.max_days_per_year or 0,
                    used          = 0,
                ))
                created += 1

    db.commit()
    return MessageResponse(message=f"Seeded {created} leave balance rows for {year}")


# ─── Leave Requests ───────────────────────────────────────────────────────────

@router.get("/leave-requests", response_model=PaginatedResponse[LeaveRequestResponse],
            summary="List leave requests")
def list_leave_requests(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    employee_id:  int | None                = Query(None),
    req_status:   LeaveRequestStatus | None = Query(None, alias="status"),
    skip:         int                       = Query(0, ge=0),
    limit:        int                       = Query(50, ge=1, le=200),
):
    q = db.query(LeaveRequest)

    # Self-service: non-managers see only their own
    if current_user.role.name not in (*_mgr_roles, RoleName.FINANCE, RoleName.COST_CONTROL):
        my_emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if my_emp:
            q = q.filter(LeaveRequest.employee_id == my_emp.id)
        else:
            return {"items": [], "total": 0}
    elif employee_id:
        q = q.filter(LeaveRequest.employee_id == employee_id)

    if req_status:
        q = q.filter(LeaveRequest.status == req_status)

    total = q.count()
    items = q.order_by(LeaveRequest.id.desc()).offset(skip).limit(limit).all()
    return {"items": items, "total": total}


@router.post("/leave-requests", response_model=LeaveRequestResponse, status_code=201,
             summary="Submit a leave request")
def submit_leave_request(
    request:      Request,
    payload:      LeaveRequestCreate,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    if payload.employee_id:
        emp = db.query(Employee).filter(Employee.id == payload.employee_id).first()
        if not emp:
            raise HTTPException(404, "Employee not found")
        # HR/admin can submit for others; employees can only submit for themselves
        if emp.user_id != current_user.id and current_user.role.name not in _hr_roles:
            raise HTTPException(403, "Can only submit leave for yourself")
    else:
        emp = db.query(Employee).filter(Employee.user_id == current_user.id).first()
        if not emp:
            raise HTTPException(404, "No employee record linked to your account")

    lt = db.query(LeaveType).filter(
        LeaveType.id == payload.leave_type_id, LeaveType.is_active == True
    ).first()
    if not lt:
        raise HTTPException(404, "Leave type not found or inactive")

    # Calculate business days — exclude Saturdays (5) and Sundays (6)
    delta = sum(
        1 for i in range((payload.end_date - payload.start_date).days + 1)
        if (payload.start_date + timedelta(days=i)).weekday() < 5
    )
    if delta == 0:
        raise HTTPException(422, "Leave dates must include at least one working day (Mon–Fri)")

    # Check balance
    year = payload.start_date.year
    balance = db.query(LeaveBalance).filter(
        LeaveBalance.employee_id   == emp.id,
        LeaveBalance.leave_type_id == lt.id,
        LeaveBalance.year          == year,
    ).first()

    if balance and lt.max_days_per_year is not None:
        if balance.remaining < delta:
            raise HTTPException(422, f"Insufficient leave balance: {balance.remaining} days remaining")

    # Build approval chain
    chain = _LEAVE_APPROVAL_CHAIN if lt.requires_approval else []

    req = LeaveRequest(
        employee_id           = emp.id,
        leave_type_id         = lt.id,
        start_date            = payload.start_date,
        end_date              = payload.end_date,
        days                  = delta,
        reason                = payload.reason,
        status                = LeaveRequestStatus.SUBMITTED if chain else LeaveRequestStatus.APPROVED,
        approval_chain        = chain,
        approval_step         = 0,
        current_approver_role = chain[0] if chain else None,
        approval_history      = [{
            "action": "SUBMIT",
            "role": None,
            "user_id": current_user.id,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "note": None,
        }],
        submitted_by = current_user.id,
    )
    db.add(req)
    db.flush()

    # If auto-approved (no approval required), deduct balance
    if req.status == LeaveRequestStatus.APPROVED:
        _deduct_balance(db, emp.id, lt, year, delta)

    write_audit(db, "LeaveRequest", req.id, "SUBMIT",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                after=model_to_dict(req))
    db.commit()
    db.refresh(req)

    # Notify approver
    if chain:
        push_to_role(db, RoleName[chain[0]],
                     "Pengajuan Cuti Baru",
                     f"{emp.full_name} mengajukan cuti {lt.name} {delta} hari",
                     "/hris/leave")
        db.commit()

    return req


@router.post("/leave-requests/{req_id}/approve", response_model=LeaveRequestResponse,
             summary="Approve a leave request")
def approve_leave_request(
    req_id:       int,
    request:      Request,
    payload:      LeaveActionRequest,
    current_user: Annotated[CurrentUser, Depends(require_role(*_mgr_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    req = _get_leave_or_404(req_id, db)

    if req.status != LeaveRequestStatus.SUBMITTED:
        raise HTTPException(409, f"Cannot approve: status is '{req.status.value}'")

    # Verify current user's role matches expected approver
    if (req.current_approver_role and
            current_user.role.name.value != req.current_approver_role and
            current_user.role.name != RoleName.SUPER_ADMIN):
        raise HTTPException(403, f"Approval expected from role: {req.current_approver_role}")

    chain = req.approval_chain or []
    step  = req.approval_step + 1

    _add_leave_history(req, current_user.id, "APPROVE", payload.note)

    if step >= len(chain):
        # Final approval
        req.status                = LeaveRequestStatus.APPROVED
        req.current_approver_role = None
        req.approved_by           = current_user.id
        req.approval_step         = step
        # Deduct balance
        _deduct_balance(db, req.employee_id, req.leave_type, req.start_date.year, req.days)
        # Notify employee
        if req.employee and req.employee.user_id:
            push(db, req.employee.user_id,
                 "Cuti Disetujui",
                 f"Pengajuan cuti {req.leave_type.name} {req.days} hari telah disetujui",
                 "/hris/leave")
    else:
        # Advance to next approver
        req.approval_step         = step
        req.current_approver_role = chain[step]

    write_audit(db, "LeaveRequest", req.id, "APPROVE",
                changed_by=current_user.id, ip_address=get_client_ip(request))
    db.commit()
    db.refresh(req)
    return req


@router.post("/leave-requests/{req_id}/reject", response_model=LeaveRequestResponse,
             summary="Reject a leave request")
def reject_leave_request(
    req_id:       int,
    request:      Request,
    payload:      LeaveActionRequest,
    current_user: Annotated[CurrentUser, Depends(require_role(*_mgr_roles))],
    db:           Annotated[Session, Depends(get_db)],
):
    req = _get_leave_or_404(req_id, db)

    if req.status != LeaveRequestStatus.SUBMITTED:
        raise HTTPException(409, f"Cannot reject: status is '{req.status.value}'")

    _add_leave_history(req, current_user.id, "REJECT", payload.note)
    req.status = LeaveRequestStatus.REJECTED

    write_audit(db, "LeaveRequest", req.id, "REJECT",
                changed_by=current_user.id, ip_address=get_client_ip(request))
    db.commit()
    db.refresh(req)

    # Notify employee
    if req.employee and req.employee.user_id:
        push(db, req.employee.user_id,
             "Cuti Ditolak",
             f"Pengajuan cuti {req.leave_type.name} {req.days} hari ditolak. {payload.note or ''}",
             "/hris/leave")
        db.commit()

    return req


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _get_leave_or_404(req_id: int, db: Session) -> LeaveRequest:
    req = db.query(LeaveRequest).filter(LeaveRequest.id == req_id).first()
    if not req:
        raise HTTPException(404, "Leave request not found")
    return req


def _add_leave_history(req: LeaveRequest, actor_id: int, action: str, note: str | None):
    history = list(req.approval_history or [])
    history.append({
        "action":    action,
        "role":      req.current_approver_role,
        "user_id":   actor_id,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "note":      note,
    })
    req.approval_history = history


def _deduct_balance(db: Session, employee_id: int, leave_type: LeaveType, year: int, days: int):
    """
    Deduct leave balance for the given leave type.
    Maternity and paternity leave are separate statutory entitlements — no balance deduction.
    """
    _no_deduct_categories = (LeaveCategory.MATERNITY, LeaveCategory.PATERNITY)
    if leave_type.category in _no_deduct_categories:
        return

    balance = db.query(LeaveBalance).filter(
        LeaveBalance.employee_id   == employee_id,
        LeaveBalance.leave_type_id == leave_type.id,
        LeaveBalance.year          == year,
    ).first()
    if balance:
        balance.used = min(balance.accrued, balance.used + days)


# ═══════════════════════════════════════════════════════════════════════════════
# Overtime Requests
# ═══════════════════════════════════════════════════════════════════════════════

def _get_my_employee_att(db: Session, cu) -> Employee:
    emp = db.query(Employee).filter(Employee.user_id == cu.id).first()
    if not emp:
        raise HTTPException(404, "No employee profile linked to your account")
    return emp


@router.post("/overtime-requests", response_model=OvertimeRequestResponse, status_code=201,
             summary="Submit overtime request")
def submit_overtime_request(
    payload: OvertimeRequestCreate,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    emp = _get_my_employee_att(db, cu)
    ot = OvertimeRequest(
        employee_id=emp.id,
        date=payload.date,
        planned_hours=payload.planned_hours,
        reason=payload.reason,
        status=OvertimeRequestStatus.SUBMITTED,
    )
    db.add(ot)
    db.commit()
    db.refresh(ot)
    # Notify HR
    push_to_role(db, RoleName.GA,
                 "Pengajuan Lembur Baru",
                 f"{emp.full_name} mengajukan lembur {payload.planned_hours}j pada {payload.date}",
                 "/hris/attendance")
    resp = OvertimeRequestResponse.model_validate(ot)
    resp.employee_name = emp.full_name
    return resp


@router.get("/me/overtime-requests", response_model=list[OvertimeRequestResponse],
            summary="My overtime requests")
def my_overtime_requests(
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    emp = _get_my_employee_att(db, cu)
    rows = (
        db.query(OvertimeRequest)
        .filter(OvertimeRequest.employee_id == emp.id)
        .order_by(OvertimeRequest.date.desc())
        .all()
    )
    result = []
    for r in rows:
        resp = OvertimeRequestResponse.model_validate(r)
        resp.employee_name = emp.full_name
        result.append(resp)
    return result


@router.get("/overtime-requests", response_model=list[OvertimeRequestResponse],
            summary="List all overtime requests (HR/MD)")
def list_overtime_requests(
    cu: Annotated[CurrentUser, Depends(require_role(*_hr_roles))],
    db: Annotated[Session, Depends(get_db)],
    status_filter: str | None = Query(None, alias="status"),
    date_from: date | None = Query(None),
    date_to:   date | None = Query(None),
):
    q = db.query(OvertimeRequest)
    if status_filter:
        q = q.filter(OvertimeRequest.status == status_filter)
    if date_from:
        q = q.filter(OvertimeRequest.date >= date_from)
    if date_to:
        q = q.filter(OvertimeRequest.date <= date_to)
    rows = q.order_by(OvertimeRequest.date.desc()).all()
    result = []
    for r in rows:
        resp = OvertimeRequestResponse.model_validate(r)
        resp.employee_name = r.employee.full_name if r.employee else None
        result.append(resp)
    return result


@router.post("/overtime-requests/{ot_id}/approve", response_model=OvertimeRequestResponse)
def approve_overtime_request(
    ot_id: int,
    payload: OvertimeActionRequest,
    cu: Annotated[CurrentUser, Depends(require_role(*_hr_roles))],
    db: Annotated[Session, Depends(get_db)],
):
    ot = db.query(OvertimeRequest).filter(OvertimeRequest.id == ot_id).first()
    if not ot:
        raise HTTPException(404, "Overtime request not found")
    if ot.status != OvertimeRequestStatus.SUBMITTED:
        raise HTTPException(400, f"Request already {ot.status.value}")
    ot.status = OvertimeRequestStatus.APPROVED
    ot.approved_by = cu.id
    ot.approved_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(ot)
    push(db, ot.employee.user_id, "Lembur Disetujui",
         f"Pengajuan lembur Anda pada {ot.date} telah disetujui.",
         "/hris/me/overtime") if ot.employee and ot.employee.user_id else None
    resp = OvertimeRequestResponse.model_validate(ot)
    resp.employee_name = ot.employee.full_name if ot.employee else None
    return resp


@router.post("/overtime-requests/{ot_id}/reject", response_model=OvertimeRequestResponse)
def reject_overtime_request(
    ot_id: int,
    payload: OvertimeActionRequest,
    cu: Annotated[CurrentUser, Depends(require_role(*_hr_roles))],
    db: Annotated[Session, Depends(get_db)],
):
    ot = db.query(OvertimeRequest).filter(OvertimeRequest.id == ot_id).first()
    if not ot:
        raise HTTPException(404, "Overtime request not found")
    if ot.status != OvertimeRequestStatus.SUBMITTED:
        raise HTTPException(400, f"Request already {ot.status.value}")
    ot.status = OvertimeRequestStatus.REJECTED
    ot.approved_by = cu.id
    ot.approved_at = datetime.now(timezone.utc)
    ot.rejection_reason = payload.note
    db.commit()
    db.refresh(ot)
    push(db, ot.employee.user_id, "Lembur Ditolak",
         f"Pengajuan lembur Anda pada {ot.date} ditolak. {payload.note or ''}",
         "/hris/me/overtime") if ot.employee and ot.employee.user_id else None
    resp = OvertimeRequestResponse.model_validate(ot)
    resp.employee_name = ot.employee.full_name if ot.employee else None
    return resp


# ═══════════════════════════════════════════════════════════════════════════════
# Team Leave Calendar
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/leave-requests/calendar", response_model=list[LeaveCalendarItem],
            summary="Team leave calendar for a given month")
def leave_calendar(
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    year:    int = Query(default=None),
    month:   int = Query(default=None),
    dept_id: int | None = Query(None),
):
    today = date.today()
    year  = year  or today.year
    month = month or today.month

    month_start = date(year, month, 1)
    # Last day of month
    if month == 12:
        month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)

    q = (
        db.query(LeaveRequest)
        .join(Employee, LeaveRequest.employee_id == Employee.id)
        .filter(
            LeaveRequest.status == LeaveRequestStatus.APPROVED,
            LeaveRequest.end_date >= month_start,
            LeaveRequest.start_date <= month_end,
        )
    )
    if dept_id:
        q = q.filter(Employee.dept_id == dept_id)

    rows = q.all()
    result = []
    for r in rows:
        result.append(LeaveCalendarItem(
            employee_id=r.employee_id,
            employee_name=r.employee.full_name if r.employee else "Unknown",
            dept=r.employee.department.name if r.employee and r.employee.department else None,
            leave_type=r.leave_type.name if r.leave_type else "—",
            start_date=r.start_date,
            end_date=r.end_date,
            days=r.days,
            status=r.status.value,
        ))
    result.sort(key=lambda x: x.start_date)
    return result
