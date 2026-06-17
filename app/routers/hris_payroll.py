"""
GPA-ERP HRIS — Payroll router (H3)

Endpoints:
    GET/POST  /hris/payroll/periods
    POST      /hris/payroll/periods/{id}/lock
    POST      /hris/payroll/periods/{id}/calculate
    GET       /hris/payroll/runs
    PATCH     /hris/payroll/runs/{run_id}
    GET       /hris/payroll/runs/{run_id}/slip
    GET/POST  /hris/salary-components
    GET/POST  /hris/salary-assignments
    DELETE    /hris/salary-assignments/{id}
"""
from __future__ import annotations

import csv
import io
import logging
from collections import defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import func as sql_func
from sqlalchemy.orm import Session

from app.audit import write_audit
from app.database import get_db
from app.dependencies import CurrentUser, get_current_user
from sqlalchemy.orm import joinedload

from app.hris_bpjs import calculate_bpjs
from app.hris_tax import calculate_pph21_netto, calculate_pph21_gross_up, DEFAULT_PTKP
from app.models import (
    AttendanceRecord, Employee, PayrollPeriod, PayrollRun, PaySlip,
    SalaryAssignment, SalaryComponent,
    PayrollStatus, PPh21Method, SalaryComponentType, RoleName, effective_roles,
)
from app.pdf_generator import generate_payslip
from app.schemas import (
    PayrollPeriodCreate, PayrollPeriodResponse,
    PayrollRunResponse, PayrollRunAdjust,
    SalaryComponentCreate, SalaryComponentResponse,
    SalaryAssignmentCreate, SalaryAssignmentResponse,
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["HRIS Payroll"])

_HR_ROLES      = (RoleName.SUPER_ADMIN, RoleName.MD)
_FINANCE_ROLES = (RoleName.SUPER_ADMIN, RoleName.MD, RoleName.FINANCE)


def _require(cu: Employee, roles: tuple) -> None:
    if not any(r in roles for r in effective_roles(cu.role.name)):
        raise HTTPException(403, f"Requires one of: {[r.value for r in roles]}")


# ─── Salary Components ────────────────────────────────────────────────────────

@router.get("/hris/salary-components", response_model=list[SalaryComponentResponse])
def list_salary_components(
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    return db.query(SalaryComponent).order_by(SalaryComponent.component_type, SalaryComponent.code).all()


@router.post("/hris/salary-components", response_model=SalaryComponentResponse, status_code=201)
def create_salary_component(
    body: SalaryComponentCreate,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _require(cu, _HR_ROLES)
    existing = db.query(SalaryComponent).filter_by(code=body.code).first()
    if existing:
        raise HTTPException(400, f"Component code '{body.code}' already exists")
    comp = SalaryComponent(**body.model_dump())
    db.add(comp)
    db.flush()
    write_audit(db, cu.id, "CREATE", "hris_salary_components", comp.id, None, body.model_dump())
    db.commit()
    db.refresh(comp)
    return comp


# ─── Salary Assignments ───────────────────────────────────────────────────────

@router.get("/hris/salary-assignments", response_model=list[SalaryAssignmentResponse])
def list_salary_assignments(
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    employee_id: int | None = None,
):
    q = db.query(SalaryAssignment)
    if employee_id:
        q = q.filter_by(employee_id=employee_id)
    return q.order_by(SalaryAssignment.employee_id, SalaryAssignment.effective_from.desc()).all()


@router.post("/hris/salary-assignments", response_model=SalaryAssignmentResponse, status_code=201)
def create_salary_assignment(
    body: SalaryAssignmentCreate,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _require(cu, _HR_ROLES)
    emp  = db.get(Employee, body.employee_id)
    comp = db.get(SalaryComponent, body.component_id)
    if not emp:  raise HTTPException(404, "Employee not found")
    if not comp: raise HTTPException(404, "Salary component not found")

    asgn = SalaryAssignment(**body.model_dump())
    db.add(asgn)
    db.flush()
    write_audit(db, cu.id, "CREATE", "hris_salary_assignments", asgn.id, None, body.model_dump(mode="json"))
    db.commit()
    db.refresh(asgn)
    return asgn


@router.delete("/hris/salary-assignments/{asgn_id}", status_code=204)
def delete_salary_assignment(
    asgn_id: int,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _require(cu, _HR_ROLES)
    asgn = db.get(SalaryAssignment, asgn_id)
    if not asgn:
        raise HTTPException(404, "Assignment not found")
    db.delete(asgn)
    db.flush()
    write_audit(db, cu.id, "DELETE", "hris_salary_assignments", asgn_id, None, None)
    db.commit()


# ─── Payroll Periods ──────────────────────────────────────────────────────────

@router.get("/hris/payroll/periods", response_model=list[PayrollPeriodResponse])
def list_periods(
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    return db.query(PayrollPeriod).order_by(PayrollPeriod.year.desc(), PayrollPeriod.month.desc()).all()


@router.post("/hris/payroll/periods", response_model=PayrollPeriodResponse, status_code=201)
def create_period(
    body: PayrollPeriodCreate,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _require(cu, _HR_ROLES)
    if not (1 <= body.month <= 12):
        raise HTTPException(400, "month must be 1–12")
    existing = db.query(PayrollPeriod).filter_by(year=body.year, month=body.month).first()
    if existing:
        raise HTTPException(400, f"Period {body.year}-{body.month:02d} already exists")
    period = PayrollPeriod(year=body.year, month=body.month, status=PayrollStatus.OPEN)
    db.add(period)
    db.commit()
    db.refresh(period)
    write_audit(db, cu.id, "CREATE", "hris_payroll_periods", period.id, None, body.model_dump())
    return period


@router.post("/hris/payroll/periods/{period_id}/lock", response_model=PayrollPeriodResponse)
def lock_period(
    period_id: int,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _require(cu, _HR_ROLES)
    period = db.get(PayrollPeriod, period_id)
    if not period:
        raise HTTPException(404, "Period not found")
    if period.status != PayrollStatus.OPEN:
        raise HTTPException(400, f"Period is already {period.status}")
    period.status    = PayrollStatus.LOCKED
    period.locked_at = datetime.now(timezone.utc)
    period.locked_by = cu.id
    db.commit()
    db.refresh(period)
    write_audit(db, cu.id, "LOCK", "hris_payroll_periods", period.id, None, None)
    return period


# ─── Payroll Calculation ──────────────────────────────────────────────────────

def _build_salary_map(assignments: list[SalaryAssignment]) -> dict[str, Decimal]:
    """Sum salary assignments (pre-loaded) by component type."""
    result: dict[str, Decimal] = {}
    for asgn in assignments:
        comp_type = asgn.component.component_type.value
        result[comp_type] = result.get(comp_type, Decimal(0)) + asgn.amount
    return result


# Indonesian standard: 173 working hours per month (used for OT rate)
_HOURS_PER_MONTH = Decimal("173")


def _calc_ot_pay(basic_monthly: Decimal, ot_wd: Decimal, ot_we: Decimal, ot_hol: Decimal) -> Decimal:
    """
    Calculate overtime pay per Permenaker No. 2/2023.
    Weekday OT:  1st hour = 1.5×, subsequent hours = 2×
    Weekend OT:  hours 1–8 = 2×, hours 9+ = 3×
    Holiday OT:  same multiplier as weekend
    Hourly rate = basic_monthly / 173
    """
    if _HOURS_PER_MONTH == 0 or basic_monthly <= 0:
        return Decimal(0)
    hourly = basic_monthly / _HOURS_PER_MONTH

    # Weekday OT
    wd_pay = Decimal(0)
    if ot_wd > 0:
        first_hour = min(ot_wd, Decimal("1"))
        rest       = max(Decimal("0"), ot_wd - Decimal("1"))
        wd_pay = hourly * Decimal("1.5") * first_hour + hourly * Decimal("2") * rest

    # Weekend OT
    we_pay = Decimal(0)
    if ot_we > 0:
        first_eight = min(ot_we, Decimal("8"))
        beyond_eight = max(Decimal("0"), ot_we - Decimal("8"))
        we_pay = hourly * Decimal("2") * first_eight + hourly * Decimal("3") * beyond_eight

    # Holiday OT (same multipliers as weekend)
    hol_pay = Decimal(0)
    if ot_hol > 0:
        first_eight = min(ot_hol, Decimal("8"))
        beyond_eight = max(Decimal("0"), ot_hol - Decimal("8"))
        hol_pay = hourly * Decimal("2") * first_eight + hourly * Decimal("3") * beyond_eight

    total = (wd_pay + we_pay + hol_pay).quantize(Decimal("1"))
    return total


@router.post("/hris/payroll/periods/{period_id}/calculate", response_model=list[PayrollRunResponse])
def calculate_period(
    period_id: int,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    pph21_method: PPh21Method = PPh21Method.NETTO,
    include_thr:  bool = False,
):
    """Run payroll for all active employees. Idempotent — recalculates existing runs."""
    _require(cu, _HR_ROLES)
    period = db.get(PayrollPeriod, period_id)
    if not period:
        raise HTTPException(404, "Period not found")
    if period.status == PayrollStatus.LOCKED:
        raise HTTPException(400, "Period is LOCKED — unlock it first or post it")
    if period.status == PayrollStatus.POSTED:
        raise HTTPException(400, "Period is already posted — cannot recalculate")

    as_of = date(period.year, period.month, 1)
    period_start = date(period.year, period.month, 1)
    period_end   = date(period.year, period.month + 1, 1) if period.month < 12 else date(period.year + 1, 1, 1)

    employees = (
        db.query(Employee)
        .filter(Employee.status.in_(["active", "probation"]))
        .all()
    )

    # ── Pre-load all salary assignments for this period in ONE query (N+1 fix) ──
    emp_ids = [e.id for e in employees]
    all_assignments = (
        db.query(SalaryAssignment)
        .options(joinedload(SalaryAssignment.component))
        .filter(SalaryAssignment.employee_id.in_(emp_ids))
        .filter(SalaryAssignment.effective_from <= as_of)
        .filter(
            (SalaryAssignment.effective_to == None) |  # noqa: E711
            (SalaryAssignment.effective_to >= as_of)
        )
        .all()
    )
    # Group by employee_id
    asgn_by_emp: dict[int, list[SalaryAssignment]] = defaultdict(list)
    for a in all_assignments:
        asgn_by_emp[a.employee_id].append(a)

    # ── Pre-load attendance OT totals for this period in ONE query ──────────────
    ot_rows = (
        db.query(
            AttendanceRecord.employee_id,
            sql_func.sum(AttendanceRecord.hours_overtime_weekday).label("ot_wd"),
            sql_func.sum(AttendanceRecord.hours_overtime_weekend).label("ot_we"),
            sql_func.sum(AttendanceRecord.hours_overtime_holiday).label("ot_hol"),
        )
        .filter(AttendanceRecord.employee_id.in_(emp_ids))
        .filter(AttendanceRecord.date >= period_start)
        .filter(AttendanceRecord.date <  period_end)
        .group_by(AttendanceRecord.employee_id)
        .all()
    )
    ot_by_emp: dict[int, tuple[Decimal, Decimal, Decimal]] = {
        row.employee_id: (
            Decimal(str(row.ot_wd  or 0)),
            Decimal(str(row.ot_we  or 0)),
            Decimal(str(row.ot_hol or 0)),
        )
        for row in ot_rows
    }

    runs: list[PayrollRun] = []
    skipped_no_salary: list[str] = []

    for emp in employees:
        salary_map = _build_salary_map(asgn_by_emp[emp.id])

        if not salary_map:
            skipped_no_salary.append(emp.employee_no)
            logger.warning("No salary assignments for employee %s (%s) — skipping", emp.employee_no, emp.full_name)
            continue

        # Gross = BASIC + ALLOWANCE − DEDUCTION
        gross = Decimal(0)
        snapshot: dict = {}
        for comp_type_val in [SalaryComponentType.BASIC.value, SalaryComponentType.ALLOWANCE.value]:
            amt = salary_map.get(comp_type_val, Decimal(0))
            gross += amt
            snapshot[comp_type_val] = float(amt)
        deductions = salary_map.get(SalaryComponentType.DEDUCTION.value, Decimal(0))
        gross -= deductions
        snapshot["DEDUCTION"] = float(deductions)
        gross = max(Decimal(0), gross)

        # OT pay (Permenaker 2023) based on basic salary rate
        basic = salary_map.get(SalaryComponentType.BASIC.value, Decimal(0))
        ot_wd, ot_we, ot_hol = ot_by_emp.get(emp.id, (Decimal(0), Decimal(0), Decimal(0)))
        ot_pay = _calc_ot_pay(basic, ot_wd, ot_we, ot_hol)
        gross += ot_pay
        snapshot["overtime_pay"] = float(ot_pay)

        # BPJS (calculated on gross including OT)
        bpjs = calculate_bpjs(gross)
        bpjs_tk_emp  = bpjs["jht_employee"] + bpjs["jp_employee"]
        bpjs_tk_er   = bpjs["jht_employer"] + bpjs["jp_employer"] + bpjs["jkk_employer"] + bpjs["jkm_employer"]
        bpjs_kes_emp = bpjs["kes_employee"]
        bpjs_kes_er  = bpjs["kes_employer"]

        # PPh 21 — use employee's PTKP status and mid-year projection
        ptkp_status = emp.ptkp_status or DEFAULT_PTKP
        months_remaining = 12
        if emp.join_date and emp.join_date.year == period.year:
            months_remaining = max(1, 12 - emp.join_date.month + 1)

        taxable = gross - bpjs_tk_emp - bpjs_kes_emp
        if pph21_method == PPh21Method.NETTO:
            pph21     = calculate_pph21_netto(taxable, ptkp_status, months_remaining)
            tunjangan = Decimal(0)
        else:
            tunjangan, pph21 = calculate_pph21_gross_up(taxable, ptkp_status, months_remaining)

        # THR (pro-rata or full basic)
        thr = None
        if include_thr and emp.join_date:
            months_worked = (as_of.year - emp.join_date.year) * 12 + (as_of.month - emp.join_date.month)
            if months_worked >= 12:
                thr = basic
            elif months_worked > 0:
                thr = (basic * months_worked / 12).quantize(Decimal("1"))

        # Net pay
        net = gross + tunjangan - bpjs_tk_emp - bpjs_kes_emp - pph21
        net = max(Decimal(0), net)

        snapshot.update({
            "bpjs_jht_employee": float(bpjs["jht_employee"]),
            "bpjs_jp_employee":  float(bpjs["jp_employee"]),
            "bpjs_kes_employee": float(bpjs_kes_emp),
            "pph21":             float(pph21),
            "tunjangan_pajak":   float(tunjangan),
            "ptkp_status":       ptkp_status,
            "months_remaining":  months_remaining,
        })

        # Upsert
        run = db.query(PayrollRun).filter_by(period_id=period.id, employee_id=emp.id).first()
        if run is None:
            run = PayrollRun(period_id=period.id, employee_id=emp.id)
            db.add(run)

        run.gross_salary        = gross
        run.bpjs_tk_employee    = bpjs_tk_emp
        run.bpjs_tk_employer    = bpjs_tk_er
        run.bpjs_kes_employee   = bpjs_kes_emp
        run.bpjs_kes_employer   = bpjs_kes_er
        run.pph21_amount        = pph21
        run.pph21_method        = pph21_method
        run.net_salary          = net
        run.thr_amount          = thr
        run.components_snapshot = snapshot
        runs.append(run)

    write_audit(db, cu.id, "CALCULATE", "hris_payroll_periods", period.id, None,
                {"employee_count": len(runs), "skipped_no_salary": skipped_no_salary})
    db.commit()
    for r in runs:
        db.refresh(r)

    logger.info("Payroll calculated: period=%s, employees=%s, skipped=%s",
                period_id, len(runs), len(skipped_no_salary))
    return runs


# ─── Payroll Runs ─────────────────────────────────────────────────────────────

@router.get("/hris/payroll/runs", response_model=list[PayrollRunResponse])
def list_runs(
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    period_id:   int | None = None,
    employee_id: int | None = None,
):
    _require(cu, _HR_ROLES + _FINANCE_ROLES)
    q = db.query(PayrollRun)
    if period_id:   q = q.filter_by(period_id=period_id)
    if employee_id: q = q.filter_by(employee_id=employee_id)
    return q.order_by(PayrollRun.period_id.desc(), PayrollRun.employee_id).all()


@router.patch("/hris/payroll/runs/{run_id}", response_model=PayrollRunResponse)
def adjust_run(
    run_id: int,
    body: PayrollRunAdjust,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    _require(cu, _HR_ROLES + _FINANCE_ROLES)
    run = db.get(PayrollRun, run_id)
    if not run:
        raise HTTPException(404, "Payroll run not found")
    period = db.get(PayrollPeriod, run.period_id)
    if period and period.status == PayrollStatus.POSTED:
        raise HTTPException(400, "Cannot adjust a posted period")

    before = {
        "gross_salary":   float(run.gross_salary),
        "thr_amount":     float(run.thr_amount) if run.thr_amount else None,
        "pph21_method":   run.pph21_method.value,
        "cost_centre_id": run.cost_centre_id,
    }

    if body.gross_salary   is not None: run.gross_salary   = body.gross_salary
    if body.thr_amount     is not None: run.thr_amount     = body.thr_amount
    if body.pph21_method   is not None: run.pph21_method   = body.pph21_method
    if body.cost_centre_id is not None: run.cost_centre_id = body.cost_centre_id

    # Recalculate net
    bpjs     = calculate_bpjs(run.gross_salary)
    bpjs_emp = bpjs["jht_employee"] + bpjs["jp_employee"] + bpjs["kes_employee"]
    if run.pph21_method == PPh21Method.NETTO:
        pph21     = calculate_pph21_netto(run.gross_salary - bpjs_emp)
        tunjangan = Decimal(0)
    else:
        tunjangan, pph21 = calculate_pph21_gross_up(run.gross_salary - bpjs_emp)
    run.pph21_amount = pph21
    run.net_salary   = run.gross_salary + tunjangan - bpjs_emp - pph21

    db.commit()
    db.refresh(run)
    write_audit(db, cu.id, "ADJUST", "hris_payroll_runs", run.id, before, body.model_dump(mode="json", exclude_none=True))
    return run


# ─── Pay Slip ─────────────────────────────────────────────────────────────────

def _can_view_payslip(cu: "CurrentUser", run: PayrollRun, db: Session) -> bool:
    """Payroll roles (SA/MD/FINANCE) can view any slip; everyone else only their own."""
    if cu.role.name in _HR_ROLES + _FINANCE_ROLES:
        return True
    emp = db.query(Employee).filter(Employee.user_id == cu.id).first()
    return emp is not None and emp.id == run.employee_id


@router.get("/hris/payroll/runs/{run_id}/slip")
def get_payslip(
    run_id: int,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    """Return a structured pay slip JSON for a payroll run."""
    run = db.get(PayrollRun, run_id)
    if not run:
        raise HTTPException(404, "Payroll run not found")
    if not _can_view_payslip(cu, run, db):
        raise HTTPException(403, "You can only view your own pay slip")
    emp    = db.get(Employee, run.employee_id)
    period = db.get(PayrollPeriod, run.period_id)
    if not emp or not period:
        raise HTTPException(404, "Employee or period record not found — data may be orphaned")

    return {
        "period":            f"{period.year}-{period.month:02d}",
        "employee_no":       emp.employee_no   if emp else None,
        "employee_name":     emp.full_name     if emp else None,
        "department":        emp.department.name if emp and emp.department else None,
        "gross_salary":      float(run.gross_salary),
        "bpjs_tk_employee":  float(run.bpjs_tk_employee),
        "bpjs_tk_employer":  float(run.bpjs_tk_employer),
        "bpjs_kes_employee": float(run.bpjs_kes_employee),
        "bpjs_kes_employer": float(run.bpjs_kes_employer),
        "pph21_amount":      float(run.pph21_amount),
        "pph21_method":      run.pph21_method.value,
        "thr_amount":        float(run.thr_amount) if run.thr_amount else None,
        "net_salary":        float(run.net_salary),
        "components":        run.components_snapshot,
    }


@router.get("/hris/payroll/runs/{run_id}/slip.pdf")
def download_payslip_pdf(
    run_id: int,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    """Generate and return a PDF pay slip for the given payroll run."""
    run = db.get(PayrollRun, run_id)
    if not run:
        raise HTTPException(404, "Payroll run not found")
    if not _can_view_payslip(cu, run, db):
        raise HTTPException(403, "You can only view your own pay slip")

    emp    = db.get(Employee, run.employee_id)
    period = db.get(PayrollPeriod, run.period_id)
    if not emp or not period:
        raise HTTPException(404, "Employee or period not found")

    MONTHS_ID = [
        "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember",
    ]
    period_label = f"{MONTHS_ID[period.month]} {period.year}"

    snapshot      = run.components_snapshot or {}
    tunjangan_pajak = Decimal(str(snapshot.get("tunjangan_pajak", 0)))

    pdf_bytes = generate_payslip(
        period_label      = period_label,
        employee_no       = emp.employee_no,
        employee_name     = emp.full_name,
        department        = emp.department.name if emp.department else None,
        bank_name         = emp.bank_name,
        bank_account      = emp.bank_account,
        gross_salary      = run.gross_salary,
        bpjs_tk_employee  = run.bpjs_tk_employee,
        bpjs_kes_employee = run.bpjs_kes_employee,
        bpjs_tk_employer  = run.bpjs_tk_employer,
        bpjs_kes_employer = run.bpjs_kes_employer,
        pph21_amount      = run.pph21_amount,
        pph21_method      = run.pph21_method.value,
        tunjangan_pajak   = tunjangan_pajak,
        thr_amount        = run.thr_amount,
        net_salary        = run.net_salary,
        generated_at      = datetime.now(timezone.utc),
        components_snapshot = run.components_snapshot,
    )

    filename = f"slip-gaji-{emp.employee_no}-{period.year}{period.month:02d}.pdf"
    return Response(
        content     = pdf_bytes,
        media_type  = "application/pdf",
        headers     = {"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Period Approve / Post ────────────────────────────────────────────────────

@router.post("/hris/payroll/periods/{period_id}/post", response_model=PayrollPeriodResponse)
def post_period(
    period_id: int,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    """
    Approve and post a locked payroll period.
    Sets status LOCKED → POSTED.
    Requires MD or Finance role.
    """
    _require(cu, _FINANCE_ROLES)
    period = db.get(PayrollPeriod, period_id)
    if not period:
        raise HTTPException(404, "Period not found")
    if period.status != PayrollStatus.LOCKED:
        raise HTTPException(
            400,
            f"Period must be LOCKED before posting (current: {period.status.value}). "
            "Lock the period first after calculating payroll.",
        )

    runs = db.query(PayrollRun).filter_by(period_id=period.id).all()
    if not runs:
        raise HTTPException(400, "No payroll runs found — run calculation first")

    period.status    = PayrollStatus.POSTED
    period.locked_by = cu.id   # reuse locked_by as poster (approved_by)

    # ── Create ERP Expense records for each run that has a cost_centre_id ────
    from app.models import Expense, ExpenseStatus, ExpenseType, CostCode
    MONTH_ID = ["","Januari","Februari","Maret","April","Mei","Juni",
                "Juli","Agustus","September","Oktober","November","Desember"]
    period_label = f"{MONTH_ID[period.month]} {period.year}"
    # Look up PERSONNEL cost code (fallback: any code with category PERSONNEL or first active)
    from app.models import CostCodeCategory
    personnel_cc = (
        db.query(CostCode)
        .filter(CostCode.category == CostCodeCategory.PERSONNEL, CostCode.is_active == True)
        .first()
    )
    expenses_created = 0
    for run in runs:
        if run.expense_id:
            continue  # already posted
        if personnel_cc is None:
            continue  # no suitable cost code — skip silently
        emp = db.get(Employee, run.employee_id)
        exp = Expense(
            expense_type=ExpenseType.REGULAR,
            cost_code_id=personnel_cc.id,
            cost_centre_id=run.cost_centre_id,
            amount=run.net_salary,
            description=f"Penggajian {period_label} — {emp.full_name if emp else run.employee_id}",
            status=ExpenseStatus.APPROVED,
            submitted_by=cu.id,
            approved_by=cu.id,
        )
        db.add(exp)
        db.flush()
        run.expense_id = exp.id
        expenses_created += 1

    db.commit()
    db.refresh(period)

    write_audit(
        db, cu.id, "POST", "hris_payroll_periods", period.id,
        {"status": "LOCKED"},
        {"status": "POSTED", "employee_count": len(runs), "expenses_created": expenses_created},
    )
    logger.info(f"Payroll posted: period={period_id}, runs={len(runs)}, expenses_created={expenses_created}, by user={cu.id}")
    return period


# ─── Bank CSV Export ──────────────────────────────────────────────────────────

@router.get("/hris/payroll/periods/{period_id}/export/bank")
def export_bank_csv(
    period_id: int,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    bank: str = "BCA",
):
    """
    Export payroll disbursement as a bank transfer CSV.
    Supports format: BCA (default), MANDIRI, BNI, BRI.
    Returns a CSV file download.
    """
    period = db.get(PayrollPeriod, period_id)
    if not period:
        raise HTTPException(404, "Period not found")
    if period.status not in (PayrollStatus.LOCKED, PayrollStatus.POSTED):
        raise HTTPException(400, "Period must be LOCKED or POSTED to export")

    runs = (
        db.query(PayrollRun)
        .options(joinedload(PayrollRun.employee))
        .filter_by(period_id=period.id)
        .order_by(PayrollRun.employee_id)
        .all()
    )

    MONTHS_ID = [
        "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember",
    ]
    period_label = f"{MONTHS_ID[period.month]} {period.year}"
    bank_up      = bank.upper()

    buf = io.StringIO()
    writer = csv.writer(buf)

    # ── Header row (varies by bank convention) ──────────────────────
    if bank_up == "MANDIRI":
        writer.writerow([
            "NO", "NAMA_PENERIMA", "NO_REKENING", "NAMA_BANK",
            "NOMINAL", "KETERANGAN",
        ])
    elif bank_up in ("BNI", "BRI"):
        writer.writerow([
            "NO", "NAMA", "REKENING", "BANK", "JUMLAH", "KETERANGAN",
        ])
    else:  # BCA default
        writer.writerow([
            "NO", "NAMA PENERIMA", "NO REKENING", "NOMINAL",
            "KETERANGAN",
        ])

    # ── Data rows ───────────────────────────────────────────────────
    for idx, run in enumerate(runs, start=1):
        emp  = run.employee
        net  = round(float(run.net_salary))
        desc = f"Gaji {period_label} - {emp.full_name if emp else str(run.employee_id)}"

        if bank_up in ("MANDIRI", "BNI", "BRI"):
            writer.writerow([
                idx,
                emp.full_name if emp else "",
                emp.bank_account if emp else "",
                emp.bank_name if emp else "",
                net,
                desc,
            ])
        else:  # BCA
            writer.writerow([
                idx,
                emp.full_name if emp else "",
                emp.bank_account if emp else "",
                net,
                desc,
            ])

    # ── Total row ───────────────────────────────────────────────────
    total = round(sum(float(r.net_salary) for r in runs))
    if bank_up in ("MANDIRI", "BNI", "BRI"):
        writer.writerow(["", "TOTAL", "", "", total, ""])
    else:
        writer.writerow(["", "TOTAL", "", total, ""])

    csv_bytes = buf.getvalue().encode("utf-8-sig")   # BOM for Excel compatibility
    filename  = f"payroll-{period.year}{period.month:02d}-{bank_up}.csv"

    return Response(
        content    = csv_bytes,
        media_type = "text/csv; charset=utf-8",
        headers    = {"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── BPJS Monthly Report Export ───────────────────────────────────────────────

@router.get("/hris/payroll/periods/{period_id}/export/bpjs")
def export_bpjs_report(
    period_id: int,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    """Export BPJS monthly contribution report as Excel (2 sheets: TK & Kes)."""
    _require(cu, _FINANCE_ROLES)
    period = db.get(PayrollPeriod, period_id)
    if not period:
        raise HTTPException(404, "Period not found")
    if period.status not in (PayrollStatus.LOCKED, PayrollStatus.POSTED):
        raise HTTPException(400, "Period must be LOCKED or POSTED to export BPJS report")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = openpyxl.Workbook()
    MONTH_ID = ["","Januari","Februari","Maret","April","Mei","Juni",
                "Juli","Agustus","September","Oktober","November","Desember"]
    period_label = f"{MONTH_ID[period.month]} {period.year}"

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True)

    def _mk_header(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

    # Sheet 1 — BPJS Ketenagakerjaan (TK)
    ws_tk = wb.active
    ws_tk.title = "BPJS TK"
    _mk_header(ws_tk, [
        "No", "No Karyawan", "Nama", "No BPJS TK",
        "Gaji Dasar",
        "JHT Karyawan (2%)", "JHT Perusahaan (3.7%)",
        "JP Karyawan (1%)", "JP Perusahaan (2%)",
        "JKK Perusahaan (0.89%)", "JKM Perusahaan (0.3%)",
        "Total Karyawan", "Total Perusahaan",
    ])

    runs = db.query(PayrollRun).filter_by(period_id=period.id).all()
    for idx, run in enumerate(runs, 1):
        emp = db.get(Employee, run.employee_id)
        snap = run.components_snapshot or {}
        bpjs = snap.get("bpjs", {})
        ws_tk.append([
            idx,
            emp.employee_no if emp else "",
            emp.full_name if emp else "",
            emp.bpjs_tk_no if emp else "",
            float(run.gross_salary),
            float(bpjs.get("jht_employee", run.bpjs_tk_employee)),
            float(bpjs.get("jht_employer", run.bpjs_tk_employer)),
            float(bpjs.get("jp_employee", 0)),
            float(bpjs.get("jp_employer", 0)),
            float(bpjs.get("jkk_employer", 0)),
            float(bpjs.get("jkm_employer", 0)),
            float(run.bpjs_tk_employee),
            float(run.bpjs_tk_employer),
        ])

    # Sheet 2 — BPJS Kesehatan (Kes)
    ws_kes = wb.create_sheet("BPJS Kes")
    _mk_header(ws_kes, [
        "No", "No Karyawan", "Nama", "No BPJS Kes",
        "Gaji Dasar",
        "Kes Karyawan (1%)", "Kes Perusahaan (4%)",
        "Total Iuran",
    ])
    for idx, run in enumerate(runs, 1):
        emp = db.get(Employee, run.employee_id)
        total_kes = float(run.bpjs_kes_employee) + float(run.bpjs_kes_employer)
        ws_kes.append([
            idx,
            emp.employee_no if emp else "",
            emp.full_name if emp else "",
            emp.bpjs_kes_no if emp else "",
            float(run.gross_salary),
            float(run.bpjs_kes_employee),
            float(run.bpjs_kes_employer),
            total_kes,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"BPJS-{period.year}{period.month:02d}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── PPh 21 Form 1721-A1 Export ──────────────────────────────────────────────

@router.get("/hris/payroll/employees/{employee_id}/form-1721")
def export_form_1721(
    employee_id: int,
    cu: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    year: int = Query(default=None),
):
    """
    Export PPh 21 annual summary (Form 1721-A1) for an employee.
    Returns Excel with monthly income breakdown and annual tax reconciliation.
    """
    _require(cu, _FINANCE_ROLES)
    from datetime import date as dt_date
    year = year or dt_date.today().year
    emp = db.get(Employee, employee_id)
    if not emp:
        raise HTTPException(404, "Employee not found")

    # All payroll runs for this employee in the given year
    runs_with_period = (
        db.query(PayrollRun, PayrollPeriod)
        .join(PayrollPeriod, PayrollRun.period_id == PayrollPeriod.id)
        .filter(PayrollRun.employee_id == employee_id, PayrollPeriod.year == year)
        .order_by(PayrollPeriod.month)
        .all()
    )

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"1721-A1 {year}"

    MONTH_ID = ["","Januari","Februari","Maret","April","Mei","Juni",
                "Juli","Agustus","September","Oktober","November","Desember"]

    # Header
    ws.merge_cells("A1:J1")
    ws["A1"] = f"FORM 1721-A1 — BUKTI PEMOTONGAN PPh PASAL 21"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Tahun Pajak: {year}"
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])

    # Employee info
    ws.append(["Nama Karyawan",   ":", emp.full_name])
    ws.append(["No. Karyawan",    ":", emp.employee_no])
    ws.append(["NIK",             ":", emp.nik or "—"])
    ws.append(["NPWP",            ":", emp.npwp or "—"])
    ws.append(["Status PTKP",     ":", emp.ptkp_status or "TK/0"])
    ws.append([])

    # Monthly breakdown header
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True)
    headers = ["Bulan", "Gaji Bruto", "BPJS TK Pegawai", "BPJS Kes Pegawai",
               "PPh 21 Dipotong", "Gaji Neto", "Metode PPh"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    total_gross = 0.0
    total_pph   = 0.0
    total_net   = 0.0

    month_data: dict[int, PayrollRun] = {p.month: r for r, p in runs_with_period}
    for m in range(1, 13):
        run = month_data.get(m)
        if run:
            ws.append([
                MONTH_ID[m],
                float(run.gross_salary),
                float(run.bpjs_tk_employee),
                float(run.bpjs_kes_employee),
                float(run.pph21_amount),
                float(run.net_salary),
                run.pph21_method.value,
            ])
            total_gross += float(run.gross_salary)
            total_pph   += float(run.pph21_amount)
            total_net   += float(run.net_salary)
        else:
            ws.append([MONTH_ID[m], 0, 0, 0, 0, 0, "—"])

    ws.append([])
    ws.append(["TOTAL", total_gross, "", "", total_pph, total_net, ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    ws.append([])
    ws.append(["Catatan: Dokumen ini dihasilkan secara otomatis oleh sistem GPA ERP."])

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"1721-A1-{emp.employee_no}-{year}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
