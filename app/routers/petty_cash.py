from __future__ import annotations

import io
from datetime import datetime, timezone
from decimal import Decimal
from typing import Annotated

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.audit import model_to_dict, write_audit
from app.database import get_db
from app.dependencies import CurrentUser, get_client_ip, require_role
from app.models import (
    CostCentre, CostCode, Expense, ExpenseStatus, PettyCashReport,
    PettyCashReportLine, PettyCashReportStatus, Project, RoleName, User,
)

# STAFF cannot create or post petty cash reports — GA and above only
_pc_roles = (RoleName.GA, RoleName.COST_CONTROL, RoleName.PM, RoleName.FINANCE, RoleName.MD, RoleName.SUPER_ADMIN)
PCWrite = Annotated[User, Depends(require_role(*_pc_roles))]
from app.schemas import (
    PettyCashReportCreate, PettyCashReportResponse, PettyCashReportUpdate,
)

router = APIRouter(prefix="/petty-cash-reports", tags=["Spending - Petty Cash"])


def _get_or_404(report_id: int, db: Session) -> PettyCashReport:
    report = db.query(PettyCashReport).filter(PettyCashReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Petty cash report not found")
    return report


def _validate_master_data(
    db: Session,
    project_id: int,
    cost_code_id: int,
    cost_centre_id: int | None,
) -> None:
    if not db.query(Project).filter(Project.id == project_id).first():
        raise HTTPException(status_code=404, detail="Project not found")
    if not db.query(CostCode).filter(CostCode.id == cost_code_id, CostCode.is_active == True).first():
        raise HTTPException(status_code=404, detail="Cost code not found or inactive")
    if cost_centre_id and not db.query(CostCentre).filter(
        CostCentre.id == cost_centre_id, CostCentre.is_active == True
    ).first():
        raise HTTPException(status_code=404, detail="Cost centre not found or inactive")


def _expense_description(report: PettyCashReport, line) -> str:
    date_part = f" {line.spent_on.isoformat()} -" if line.spent_on else ""
    return f"[Petty Cash {report.month}]{date_part} {line.description.strip()}"


def _recalculate_total(report: PettyCashReport) -> None:
    report.total_amount = sum((line.amount for line in report.lines), Decimal("0"))


def _create_expense_for_line(
    report: PettyCashReport,
    line: PettyCashReportLine,
    current_user: CurrentUser,
) -> Expense:
    return Expense(
        project_id=report.project_id,
        cost_code_id=report.cost_code_id,
        cost_centre_id=report.cost_centre_id,
        petty_cash_line=line,
        amount=line.amount,
        description=_expense_description(report, line),
        receipt_url=line.receipt_url,
        status=ExpenseStatus.DRAFT,
        submitted_by=current_user.id,
        approval_chain=[],
        approval_history=[],
        approval_step=0,
    )


@router.get("", response_model=list[PettyCashReportResponse], summary="List petty cash reports")
def list_reports(
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
    project_id: int | None = None,
    month: str | None = None,
    report_status: PettyCashReportStatus | None = Query(None, alias="status"),
    skip: int = 0,
    limit: int = 100,
):
    q = db.query(PettyCashReport)
    if project_id:
        q = q.filter(PettyCashReport.project_id == project_id)
    if month:
        q = q.filter(PettyCashReport.month == month)
    if report_status:
        q = q.filter(PettyCashReport.status == report_status)
    return q.order_by(PettyCashReport.id.desc()).offset(skip).limit(limit).all()


_PC_EXPORT_ROLES = (RoleName.GA, RoleName.FINANCE, RoleName.SUPER_ADMIN)
_PC_EXPORT_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="1E293B")
_PC_EXPORT_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_PC_EXPORT_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


@router.get("/export", summary="Export petty cash report lines to XLSX")
def export_petty_cash(
    current_user: Annotated[object, Depends(require_role(*_PC_EXPORT_ROLES))],
    db:           Annotated[Session, Depends(get_db)],
    report_id:    int | None = Query(None),
    date_from:    str | None = Query(None),
    date_to:      str | None = Query(None),
):
    q = (
        db.query(PettyCashReport)
        .options(
            joinedload(PettyCashReport.lines),
        )
    )
    if report_id:
        q = q.filter(PettyCashReport.id == report_id)
    if date_from:
        try:
            q = q.filter(PettyCashReport.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format, use ISO 8601")
    if date_to:
        try:
            q = q.filter(PettyCashReport.created_at <= datetime.fromisoformat(date_to))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format, use ISO 8601")

    reports = q.order_by(PettyCashReport.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Petty Cash"

    headers = ["No", "Tanggal", "Nama Laporan", "Keterangan", "Kategori", "Jumlah", "Status"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = _PC_EXPORT_HEADER_FILL
        cell.font      = _PC_EXPORT_HEADER_FONT
        cell.alignment = _PC_EXPORT_HEADER_ALIGN

    row_idx = 2
    seq = 1
    for report in reports:
        for line in report.lines:
            tanggal = line.spent_on.isoformat() if line.spent_on else (
                report.created_at.strftime("%Y-%m-%d") if report.created_at else ""
            )
            ws.cell(row=row_idx, column=1, value=seq)
            ws.cell(row=row_idx, column=2, value=tanggal)
            ws.cell(row=row_idx, column=3, value=report.title or report.report_no)
            ws.cell(row=row_idx, column=4, value=line.description)
            ws.cell(row=row_idx, column=5, value=report.month)
            ws.cell(row=row_idx, column=6, value=float(line.amount))
            ws.cell(row=row_idx, column=7, value=report.status.value if report.status else "")
            row_idx += 1
            seq += 1

    col_widths = [len(h) for h in headers]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                col_widths[cell.column - 1] = max(col_widths[cell.column - 1], len(str(cell.value)))
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date
    filename = f"petty-cash-{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{report_id}", response_model=PettyCashReportResponse)
def get_report(
    report_id: int,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    return _get_or_404(report_id, db)


@router.post("", response_model=PettyCashReportResponse, status_code=201, summary="Create petty cash report")
def create_report(
    request: Request,
    payload: PettyCashReportCreate,
    current_user: PCWrite,
    db: Annotated[Session, Depends(get_db)],
):
    _validate_master_data(db, payload.project_id, payload.cost_code_id, payload.cost_centre_id)

    report = PettyCashReport(
        report_no=f"PC-{payload.month.replace('-', '')}-{current_user.id}-{int(datetime.now(timezone.utc).timestamp())}",
        month=payload.month,
        project_id=payload.project_id,
        cost_code_id=payload.cost_code_id,
        cost_centre_id=payload.cost_centre_id,
        title=payload.title,
        notes=payload.notes,
        status=PettyCashReportStatus.DRAFT,
        created_by=current_user.id,
    )
    db.add(report)
    db.flush()

    for idx, item in enumerate(payload.lines, start=1):
        line = PettyCashReportLine(
            report=report,
            line_no=idx,
            spent_on=item.spent_on,
            description=item.description,
            amount=item.amount,
            receipt_url=item.receipt_url,
            source=item.source,
            ocr_text=item.ocr_text,
        )
        db.add(line)
        db.flush()
        db.add(_create_expense_for_line(report, line, current_user))

    _recalculate_total(report)
    db.flush()
    write_audit(
        db, "PettyCashReport", report.id, "CREATE",
        changed_by=current_user.id, ip_address=get_client_ip(request),
        after=model_to_dict(report),
    )
    db.commit()
    db.refresh(report)
    return report


@router.patch("/{report_id}", response_model=PettyCashReportResponse, summary="Update draft petty cash report")
def update_report(
    report_id: int,
    request: Request,
    payload: PettyCashReportUpdate,
    current_user: CurrentUser,
    db: Annotated[Session, Depends(get_db)],
):
    report = _get_or_404(report_id, db)
    if report.status != PettyCashReportStatus.DRAFT:
        raise HTTPException(status_code=409, detail="Only draft petty cash reports can be edited")
    if report.created_by != current_user.id and current_user.role.name not in (
        RoleName.SUPER_ADMIN, RoleName.COST_CONTROL
    ):
        raise HTTPException(status_code=403, detail="Not your petty cash report")

    project_id = payload.project_id if payload.project_id is not None else report.project_id
    cost_code_id = payload.cost_code_id if payload.cost_code_id is not None else report.cost_code_id
    cost_centre_id = payload.cost_centre_id if payload.cost_centre_id is not None else report.cost_centre_id
    _validate_master_data(db, project_id, cost_code_id, cost_centre_id)

    before = model_to_dict(report)
    for field, value in payload.model_dump(exclude_unset=True, exclude={"lines"}).items():
        setattr(report, field, value)

    if payload.lines is not None:
        existing = {line.id: line for line in report.lines}
        seen: set[int] = set()
        for idx, item in enumerate(payload.lines, start=1):
            data = item.model_dump(exclude_unset=True, exclude={"id"})
            if item.id and item.id in existing:
                line = existing[item.id]
                seen.add(line.id)
                if line.expense and line.expense.status != ExpenseStatus.DRAFT:
                    raise HTTPException(status_code=409, detail="Cannot edit a line after its expense leaves draft")
                for field, value in data.items():
                    setattr(line, field, value)
            else:
                if not item.description or item.amount is None:
                    raise HTTPException(status_code=422, detail="New lines require description and amount")
                line = PettyCashReportLine(report=report, **data)
                db.add(line)
                db.flush()
                db.add(_create_expense_for_line(report, line, current_user))
            line.line_no = idx
            if line.expense:
                line.expense.project_id = report.project_id
                line.expense.cost_code_id = report.cost_code_id
                line.expense.cost_centre_id = report.cost_centre_id
                line.expense.amount = line.amount
                line.expense.description = _expense_description(report, line)
                line.expense.receipt_url = line.receipt_url

        for line in list(report.lines):
            if line.id and line.id not in seen and line.id in existing:
                if line.expense and line.expense.status != ExpenseStatus.DRAFT:
                    raise HTTPException(status_code=409, detail="Cannot remove a line after its expense leaves draft")
                if line.expense:
                    db.delete(line.expense)
                db.delete(line)

    _recalculate_total(report)
    write_audit(
        db, "PettyCashReport", report.id, "UPDATE",
        changed_by=current_user.id, ip_address=get_client_ip(request),
        before=before, after=model_to_dict(report),
    )
    db.commit()
    db.refresh(report)
    return report


@router.post("/{report_id}/post", response_model=PettyCashReportResponse, summary="Mark petty cash report posted")
def post_report(
    report_id: int,
    request: Request,
    current_user: PCWrite,
    db: Annotated[Session, Depends(get_db)],
):
    report = _get_or_404(report_id, db)
    if report.status != PettyCashReportStatus.DRAFT:
        raise HTTPException(status_code=409, detail="Only draft reports can be posted")

    before = model_to_dict(report)
    report.status = PettyCashReportStatus.POSTED
    report.posted_at = datetime.now(timezone.utc)
    write_audit(
        db, "PettyCashReport", report.id, "POST",
        changed_by=current_user.id, ip_address=get_client_ip(request),
        before=before, after=model_to_dict(report),
    )
    db.commit()
    db.refresh(report)
    return report

