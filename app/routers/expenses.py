"""
Expenses router — full lifecycle with multi-step approval matrix.

Lifecycle:
  draft  ──submit──►  submitted
  submitted  ──verify──►  verified      (COST_CONTROL)
  verified  ──approve──►  approved       (role from matrix chain)
  approved  ──pay──►  paid              (FINANCE)
  paid  ──lock──►  hard_locked          (SUPER_ADMIN / period close)
  any (non-locked)  ──reject──►  rejected  (any approver in chain / FINANCE)
  rejected  ──resubmit──►  submitted

The approval_chain is built from the ApprovalRule matrix at submit time and
stored on the Expense so it is immutable thereafter.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Annotated

import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy import func, or_

UPLOAD_DIR = Path("uploads/receipts")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
_ALLOWED_TYPES = {
    "image/jpeg", "image/jpg", "image/png", "image/webp",
    "image/heic", "image/heif", "application/pdf",
}
_MAX_SIZE = 10 * 1024 * 1024  # 10 MB
from sqlalchemy.orm import Session, joinedload

from app.audit import model_to_dict, write_audit
from app.database import get_db
from app.dependencies import (
    CurrentUser, get_client_ip, get_required_approvers_from_matrix, require_role,
)
from app.models import (
    CostCentre, CostCode, Expense, ExpenseStatus, ExpenseType, Project, RoleName,
)
from app.notify import push, push_to_role
from app.schemas import (
    ExpenseActionRequest, ExpenseCreate, ExpenseRejectRequest,
    ExpenseResponse, ExpenseStats, ExpenseUpdate, MessageResponse, PaginatedResponse,
)

router = APIRouter(prefix="/expenses", tags=["Spending – Expenses"])


def _get_or_404(expense_id: int, db: Session) -> Expense:
    e = db.query(Expense).filter(Expense.id == expense_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Expense not found")
    return e


def _add_history_event(expense: Expense, actor_id: int, action: str, note: str | None = None):
    history = list(expense.approval_history or [])
    history.append({
        "action":    action,
        "role":      expense.current_approver_role,
        "user_id":   actor_id,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "note":      note,
    })
    expense.approval_history = history


# ─── List / Get ──────────────────────────────────────────────────────────────

@router.get("", response_model=PaginatedResponse[ExpenseResponse], summary="List expenses")
def list_expenses(
    current_user:   CurrentUser,
    db:             Annotated[Session, Depends(get_db)],
    project_id:     int | None          = None,
    expense_status: ExpenseStatus | None = Query(None, alias="status"),
    my_queue:       bool                = Query(False, description="Only expenses pending MY role"),
    search:         str | None          = Query(None, description="Search description, vendor, or reference"),
    skip:           int = Query(0, ge=0),
    limit:          int = Query(100, ge=1, le=500),
):
    q = db.query(Expense)

    # STAFF and WORKER can only see their own submissions (confidentiality)
    self_service_roles = {RoleName.STAFF, RoleName.WORKER}
    if current_user.role.name in self_service_roles:
        q = q.filter(Expense.submitted_by == current_user.id)
    else:
        if project_id:
            q = q.filter(Expense.project_id == project_id)
        if my_queue:
            q = q.filter(Expense.current_approver_role == current_user.role.name.value)

    if expense_status:
        q = q.filter(Expense.status == expense_status)
    if search:
        q = q.filter(or_(
            Expense.description.ilike(f"%{search}%"),
            Expense.vendor_name.ilike(f"%{search}%"),
            Expense.reference_no.ilike(f"%{search}%"),
        ))
    total = q.count()
    items = q.order_by(Expense.id.desc()).offset(skip).limit(limit).all()
    return {"items": items, "total": total}


@router.get("/stats", response_model=ExpenseStats, summary="Spending summary statistics")
def get_expense_stats(
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
    project_id:   int | None = Query(None),
    date_from:    str | None = Query(None),
    date_to:      str | None = Query(None),
):
    logged_statuses  = {ExpenseStatus.SUBMITTED, ExpenseStatus.VERIFIED, ExpenseStatus.APPROVED, ExpenseStatus.PAID, ExpenseStatus.HARD_LOCKED}
    approved_statuses = {ExpenseStatus.APPROVED, ExpenseStatus.PAID, ExpenseStatus.HARD_LOCKED}
    paid_statuses     = {ExpenseStatus.PAID, ExpenseStatus.HARD_LOCKED}

    base = db.query(Expense)
    if project_id:
        base = base.filter(Expense.project_id == project_id)

    def _sum(statuses: set) -> Decimal:
        result = base.filter(Expense.status.in_(statuses)).with_entities(func.coalesce(func.sum(Expense.amount), 0)).scalar()
        return Decimal(str(result))

    def _count(st: ExpenseStatus) -> int:
        return base.filter(Expense.status == st).count()

    return ExpenseStats(
        total_logged   = _sum(logged_statuses),
        total_approved = _sum(approved_statuses),
        total_paid     = _sum(paid_statuses),
        count_by_status = {
            "draft":       _count(ExpenseStatus.DRAFT),
            "submitted":   _count(ExpenseStatus.SUBMITTED),
            "verified":    _count(ExpenseStatus.VERIFIED),
            "approved":    _count(ExpenseStatus.APPROVED),
            "paid":        _count(ExpenseStatus.PAID),
            "hard_locked": _count(ExpenseStatus.HARD_LOCKED),
            "rejected":    _count(ExpenseStatus.REJECTED),
        },
    )


# ─── Receipt upload ──────────────────────────────────────────────────────────

@router.post("/upload-receipt", summary="Upload a receipt file and get back a URL")
async def upload_receipt(
    current_user: CurrentUser,
    file: UploadFile = File(...),
):
    if file.content_type not in _ALLOWED_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"File type '{file.content_type}' not allowed. Use JPG, PNG, WebP, HEIC, or PDF.",
        )
    data = await file.read()
    if len(data) > _MAX_SIZE:
        raise HTTPException(status_code=400, detail="File must be under 10 MB")

    ext      = Path(file.filename or "receipt").suffix or ".bin"
    filename = f"{uuid.uuid4().hex}{ext}"
    dest     = UPLOAD_DIR / filename
    dest.write_bytes(data)

    return {"url": f"/uploads/receipts/{filename}", "filename": file.filename}


@router.get("/export", summary="Export expenses to XLSX")
def export_expenses(
    current_user:   CurrentUser,
    db:             Annotated[Session, Depends(get_db)],
    project_id:     int | None = Query(None),
    status:         str | None = Query(None),
    date_from:      str | None = Query(None),
    date_to:        str | None = Query(None),
):
    q = db.query(Expense)
    self_service_roles = {RoleName.STAFF, RoleName.WORKER}
    if current_user.role.name in self_service_roles:
        q = q.filter(Expense.submitted_by == current_user.id)
    elif project_id:
        q = q.filter(Expense.project_id == project_id)
    if status:
        try:
            q = q.filter(Expense.status == ExpenseStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")
    if date_from:
        try:
            q = q.filter(Expense.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format, use ISO 8601")
    if date_to:
        try:
            q = q.filter(Expense.created_at <= datetime.fromisoformat(date_to))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format, use ISO 8601")

    expenses = (
        q.options(
            joinedload(Expense.project),
            joinedload(Expense.cost_code),
            joinedload(Expense.submitter),
        )
        .order_by(Expense.id.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = [
        "ID", "Project Code", "Cost Code", "Category", "Description",
        "Vendor", "Amount", "Currency", "Status", "Submitted By", "Created At",
    ]

    header_fill = PatternFill(fill_type="solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, exp in enumerate(expenses, start=2):
        project_code = exp.project.code if exp.project else ""
        cost_code_name = exp.cost_code.code if exp.cost_code else ""
        category = exp.cost_code.category.value if exp.cost_code and exp.cost_code.category else ""
        currency = exp.project.currency if exp.project else "IDR"
        submitter_name = exp.submitter.full_name if exp.submitter else (str(exp.submitted_by) if exp.submitted_by else "")
        created_at_str = exp.created_at.strftime("%Y-%m-%d %H:%M:%S") if exp.created_at else ""

        ws.cell(row=row_idx, column=1,  value=exp.id)
        ws.cell(row=row_idx, column=2,  value=project_code)
        ws.cell(row=row_idx, column=3,  value=cost_code_name)
        ws.cell(row=row_idx, column=4,  value=category)
        ws.cell(row=row_idx, column=5,  value=exp.description)
        ws.cell(row=row_idx, column=6,  value=exp.vendor_name or "")
        ws.cell(row=row_idx, column=7,  value=float(exp.amount))
        ws.cell(row=row_idx, column=8,  value=currency)
        ws.cell(row=row_idx, column=9,  value=exp.status.value if exp.status else "")
        ws.cell(row=row_idx, column=10, value=submitter_name)
        ws.cell(row=row_idx, column=11, value=created_at_str)

    # Auto-size columns based on content
    col_widths: list[int] = [len(h) for h in headers]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                col_widths[cell.column - 1] = max(col_widths[cell.column - 1], len(str(cell.value)))
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"gpa-expenses-{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{expense_id}", response_model=ExpenseResponse)
def get_expense(
    expense_id:   int,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    expense = _get_or_404(expense_id, db)
    self_service_roles = {RoleName.STAFF, RoleName.WORKER}
    if current_user.role.name in self_service_roles and expense.submitted_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not your expense")
    return expense


# ─── Create (draft) ──────────────────────────────────────────────────────────

@router.post("", response_model=ExpenseResponse, status_code=201,
             summary="Create a draft expense")
def create_expense(
    request:      Request,
    payload:      ExpenseCreate,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    # STAFF / WORKER can only create reimbursements
    if current_user.role.name in (RoleName.STAFF, RoleName.WORKER):
        if payload.expense_type != ExpenseType.REIMBURSEMENT:
            raise HTTPException(403, "Staff and Workers can only submit reimbursements")

    # Auto-assign reimbursement cost code for STAFF/WORKER (code 99.10)
    # They don't pick a cost code in the UI — it's always "Personal Reimbursement"
    if payload.expense_type == ExpenseType.REIMBURSEMENT and current_user.role.name in (RoleName.STAFF, RoleName.WORKER):
        reimb_code = (
            db.query(CostCode)
            .filter(CostCode.code == "99.10", CostCode.is_active == True)
            .first()
        )
        if not reimb_code:
            raise HTTPException(500, "Reimbursement cost code (99.10) not found — run seed script")
        payload.cost_code_id = reimb_code.id

    # Project is required for regular expenses, optional for reimbursements
    if payload.project_id is not None:
        if not db.query(Project).filter(Project.id == payload.project_id).first():
            raise HTTPException(status_code=404, detail="Project not found")
    elif payload.expense_type == ExpenseType.REGULAR:
        raise HTTPException(status_code=422, detail="project_id is required for regular expenses")

    if not db.query(CostCode).filter(CostCode.id == payload.cost_code_id, CostCode.is_active == True).first():
        raise HTTPException(status_code=404, detail="Cost code not found or inactive")
    if payload.cost_centre_id and not db.query(CostCentre).filter(
        CostCentre.id == payload.cost_centre_id, CostCentre.is_active == True
    ).first():
        raise HTTPException(status_code=404, detail="Cost centre not found or inactive")

    expense = Expense(
        expense_type     = payload.expense_type,
        project_id       = payload.project_id,
        cost_code_id     = payload.cost_code_id,
        cost_centre_id   = payload.cost_centre_id,
        amount           = payload.amount,
        description      = payload.description,
        vendor_name      = payload.vendor_name,
        reference_no     = payload.reference_no,
        receipt_url      = payload.receipt_url,
        status           = ExpenseStatus.DRAFT,
        submitted_by     = current_user.id,
        approval_chain   = [],
        approval_history = [],
        approval_step    = 0,
    )
    db.add(expense)
    db.flush()

    write_audit(db, "Expense", expense.id, "CREATE",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                after=model_to_dict(expense))
    db.commit()
    db.refresh(expense)
    return expense


# ─── Update draft ────────────────────────────────────────────────────────────

@router.patch("/{expense_id}", response_model=ExpenseResponse,
              summary="Update a draft expense (draft only)")
def update_expense(
    expense_id:   int,
    request:      Request,
    payload:      ExpenseUpdate,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    expense = _get_or_404(expense_id, db)

    if expense.status != ExpenseStatus.DRAFT:
        raise HTTPException(status_code=409,
                            detail="Only draft expenses can be edited")
    if expense.submitted_by != current_user.id and current_user.role.name not in (
        RoleName.SUPER_ADMIN, RoleName.COST_CONTROL
    ):
        raise HTTPException(status_code=403, detail="Not your expense")

    if payload.cost_code_id is not None:
        cc = db.query(CostCode).filter(
            CostCode.id == payload.cost_code_id, CostCode.is_active == True
        ).first()
        if not cc:
            raise HTTPException(status_code=404, detail="Cost code not found or inactive")
    if payload.cost_centre_id is not None:
        centre = db.query(CostCentre).filter(
            CostCentre.id == payload.cost_centre_id, CostCentre.is_active == True
        ).first()
        if not centre:
            raise HTTPException(status_code=404, detail="Cost centre not found or inactive")

    before = model_to_dict(expense)
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(expense, field, value)

    write_audit(db, "Expense", expense.id, "UPDATE",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                before=before, after=model_to_dict(expense))
    db.commit()
    db.refresh(expense)
    return expense


# ─── Submit ──────────────────────────────────────────────────────────────────

@router.post("/{expense_id}/submit", response_model=ExpenseResponse,
             summary="Submit a draft expense for approval")
def submit_expense(
    expense_id:   int,
    request:      Request,
    payload:      ExpenseActionRequest,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    expense = _get_or_404(expense_id, db)

    if expense.submitted_by != current_user.id and current_user.role.name not in (
        RoleName.SUPER_ADMIN, RoleName.COST_CONTROL
    ):
        raise HTTPException(status_code=403, detail="Not your expense")

    allowed_from = {ExpenseStatus.DRAFT, ExpenseStatus.REJECTED}
    if expense.status not in allowed_from:
        raise HTTPException(status_code=409,
                            detail=f"Cannot submit from status '{expense.status}'")

    # Build approval chain
    if expense.expense_type == ExpenseType.REIMBURSEMENT:
        # Fixed reimbursement chain: GA receipt check → CC verify → Finance approve+pay
        chain = [RoleName.GA.value, RoleName.COST_CONTROL.value, RoleName.FINANCE.value]
    else:
        # Regular expense: chain from approval matrix, CC always first
        cost_code = db.query(CostCode).filter(CostCode.id == expense.cost_code_id).first()
        chain = get_required_approvers_from_matrix(db, expense.amount, cost_code.category)
        if RoleName.COST_CONTROL.value not in chain:
            chain.insert(0, RoleName.COST_CONTROL.value)

    before = model_to_dict(expense)
    expense.status               = ExpenseStatus.SUBMITTED
    expense.submitted_by         = current_user.id
    expense.approval_chain       = chain
    expense.approval_step        = 0
    expense.current_approver_role= chain[0] if chain else None
    expense.rejection_reason     = None

    _add_history_event(expense, current_user.id, "SUBMIT", payload.note)

    write_audit(db, "Expense", expense.id, "SUBMIT",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                before=before, after=model_to_dict(expense))
    submitter_name = expense.submitter.full_name if expense.submitter else str(expense.submitted_by)
    notify_role = RoleName.GA if expense.expense_type == ExpenseType.REIMBURSEMENT else RoleName.COST_CONTROL
    push_to_role(
        db, notify_role,
        "Reimbursement Baru" if expense.expense_type == ExpenseType.REIMBURSEMENT else "Pengeluaran Baru",
        f"{'Reimbursement' if expense.expense_type == ExpenseType.REIMBURSEMENT else 'Pengeluaran'} "
        f"#{expense.id} dari {submitter_name} menunggu {'verifikasi bukti' if expense.expense_type == ExpenseType.REIMBURSEMENT else 'verifikasi'}",
        link="/spending",
    )
    db.commit()
    db.refresh(expense)
    return expense


# ─── Verify (COST_CONTROL) ───────────────────────────────────────────────────

@router.post("/{expense_id}/verify", response_model=ExpenseResponse,
             summary="Verification step — GA (receipt check) or Cost Control")
def verify_expense(
    expense_id:   int,
    request:      Request,
    payload:      ExpenseActionRequest,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    """
    Handles the 'verify' action for both GA (receipt check, step 0 of reimbursement chain)
    and Cost Control (cost verification, all regular + step 1 of reimbursement chain).
    Caller's role must match `current_approver_role`.
    """
    expense = _get_or_404(expense_id, db)

    if expense.status != ExpenseStatus.SUBMITTED:
        raise HTTPException(status_code=409,
                            detail="Expense must be in 'submitted' status to verify")

    expected_role = expense.current_approver_role
    caller_role   = current_user.role.name.value

    # Only the expected role (or SUPER_ADMIN) may act
    if caller_role != expected_role and current_user.role.name != RoleName.SUPER_ADMIN:
        raise HTTPException(
            status_code=403,
            detail=f"This expense is waiting for {expected_role} to verify",
        )
    # Additional guard: only GA/CC/SA may call verify at all
    allowed_verify_roles = {RoleName.GA, RoleName.COST_CONTROL, RoleName.SUPER_ADMIN}
    if current_user.role.name not in allowed_verify_roles:
        raise HTTPException(403, "Only GA or Cost Control can perform verification")

    before = model_to_dict(expense)

    # Record who did what
    if current_user.role.name == RoleName.GA:
        expense.receipt_reviewed_by = current_user.id
        action_label = "RECEIPT_REVIEW"
        notif_msg    = f"Bukti reimburse #{expense.id} telah diverifikasi oleh GA"
    else:
        expense.verified_by = current_user.id
        action_label = "VERIFY"
        notif_msg    = f"Pengeluaran #{expense.id} telah diverifikasi oleh Cost Control"

    _add_history_event(expense, current_user.id, action_label, payload.note)
    _advance_chain(expense)

    write_audit(db, "Expense", expense.id, action_label,
                changed_by=current_user.id, ip_address=get_client_ip(request),
                before=before, after=model_to_dict(expense))
    if expense.submitted_by:
        push(db, expense.submitted_by,
             "Reimburse Diproses" if expense.expense_type == ExpenseType.REIMBURSEMENT else "Pengeluaran Diverifikasi",
             notif_msg, link="/spending")

    # Notify next approver in chain
    if expense.current_approver_role:
        try:
            next_role = RoleName(expense.current_approver_role)
            push_to_role(db, next_role,
                         "Menunggu Verifikasi Anda",
                         f"Pengeluaran #{expense.id} perlu ditindaklanjuti",
                         link="/spending")
        except ValueError:
            pass

    db.commit()
    db.refresh(expense)
    return expense


# ─── Approve (role from matrix) ──────────────────────────────────────────────

@router.post("/{expense_id}/approve", response_model=ExpenseResponse,
             summary="Approve at the current chain step")
def approve_expense(
    expense_id:   int,
    request:      Request,
    payload:      ExpenseActionRequest,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    expense = _get_or_404(expense_id, db)

    if expense.status not in {ExpenseStatus.SUBMITTED, ExpenseStatus.VERIFIED}:
        raise HTTPException(status_code=409,
                            detail="Expense must be submitted or verified to approve")

    expected_role = expense.current_approver_role
    if current_user.role.name.value != expected_role and current_user.role.name != RoleName.SUPER_ADMIN:
        raise HTTPException(
            status_code=403,
            detail=f"Current approval step requires role: {expected_role}",
        )

    before = model_to_dict(expense)
    expense.approved_by = current_user.id
    _add_history_event(expense, current_user.id, "APPROVE", payload.note)

    _advance_chain(expense)

    # Budget check only applies when expense is linked to a project
    over_budget      = None
    budget_remaining = None
    project = expense.project
    if project is not None:
        over_budget      = (project.total_committed + expense.amount) > project.total_revenue
        budget_remaining = project.total_revenue - project.total_committed - expense.amount

    write_audit(db, "Expense", expense.id, "APPROVE",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                before=before, after=model_to_dict(expense))
    if expense.submitted_by:
        push(
            db, expense.submitted_by,
            "Pengeluaran Disetujui",
            f"Pengeluaran #{expense.id} telah disetujui",
            link="/spending",
        )
    db.commit()
    db.refresh(expense)
    response = ExpenseResponse.model_validate(expense)
    response.over_budget      = over_budget
    response.budget_remaining = budget_remaining
    return response


# ─── Pay (FINANCE) ───────────────────────────────────────────────────────────

@router.post("/{expense_id}/pay", response_model=ExpenseResponse,
             summary="Mark expense as paid — FINANCE only")
def pay_expense(
    expense_id:   int,
    request:      Request,
    payload:      ExpenseActionRequest,
    current_user: Annotated[object, Depends(require_role(
        RoleName.FINANCE, RoleName.SUPER_ADMIN
    ))],
    db: Annotated[Session, Depends(get_db)],
):
    expense = _get_or_404(expense_id, db)

    if expense.status != ExpenseStatus.APPROVED:
        raise HTTPException(status_code=409,
                            detail="Expense must be fully approved before payment")

    before = model_to_dict(expense)
    expense.status  = ExpenseStatus.PAID
    expense.paid_by = current_user.id
    expense.current_approver_role = None
    _add_history_event(expense, current_user.id, "PAY", payload.note)

    write_audit(db, "Expense", expense.id, "PAY",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                before=before, after=model_to_dict(expense))
    if expense.submitted_by:
        push(
            db, expense.submitted_by,
            "Pengeluaran Dibayar",
            f"Pengeluaran #{expense.id} telah dibayarkan oleh Finance",
            link="/spending",
        )
    db.commit()
    db.refresh(expense)
    return expense


# ─── Hard Lock (period close) ────────────────────────────────────────────────

@router.post("/{expense_id}/lock", response_model=ExpenseResponse,
             summary="Hard-lock a paid expense — SUPER_ADMIN only")
def lock_expense(
    expense_id:   int,
    request:      Request,
    current_user: Annotated[object, Depends(require_role(RoleName.SUPER_ADMIN))],
    db:           Annotated[Session, Depends(get_db)],
):
    expense = _get_or_404(expense_id, db)

    if expense.status != ExpenseStatus.PAID:
        raise HTTPException(status_code=409,
                            detail="Only paid expenses can be hard-locked")

    before = model_to_dict(expense)
    expense.status = ExpenseStatus.HARD_LOCKED
    _add_history_event(expense, current_user.id, "HARD_LOCK")

    write_audit(db, "Expense", expense.id, "HARD_LOCK",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                before=before, after=model_to_dict(expense))
    db.commit()
    db.refresh(expense)
    return expense


# ─── Reject ──────────────────────────────────────────────────────────────────

@router.post("/{expense_id}/reject", response_model=ExpenseResponse,
             summary="Reject an expense — returns it to draft")
def reject_expense(
    expense_id:   int,
    request:      Request,
    payload:      ExpenseRejectRequest,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    expense = _get_or_404(expense_id, db)

    locked_statuses = {ExpenseStatus.PAID, ExpenseStatus.HARD_LOCKED, ExpenseStatus.REJECTED}
    if expense.status in locked_statuses:
        raise HTTPException(status_code=409,
                            detail=f"Cannot reject an expense with status '{expense.status}'")

    # Must be an approver in the chain or SUPER_ADMIN / FINANCE
    allowed_rejectors = set(expense.approval_chain or []) | {
        RoleName.SUPER_ADMIN.value, RoleName.FINANCE.value
    }
    if current_user.role.name.value not in allowed_rejectors:
        raise HTTPException(status_code=403, detail="You are not in the approval chain")

    before = model_to_dict(expense)
    expense.status                = ExpenseStatus.REJECTED
    expense.rejection_reason      = payload.reason
    expense.current_approver_role = None
    _add_history_event(expense, current_user.id, "REJECT", payload.reason)

    write_audit(db, "Expense", expense.id, "REJECT",
                changed_by=current_user.id, ip_address=get_client_ip(request),
                before=before, after=model_to_dict(expense))
    if expense.submitted_by:
        push(
            db, expense.submitted_by,
            "Pengeluaran Ditolak",
            f"Pengeluaran #{expense.id} ditolak: {expense.rejection_reason}",
            link="/spending",
        )
    db.commit()
    db.refresh(expense)
    return expense


# ─── Audit trail for one expense ─────────────────────────────────────────────

@router.get("/{expense_id}/audit", response_model=list[dict],
            summary="Full audit trail for one expense")
def expense_audit_trail(
    expense_id:   int,
    current_user: CurrentUser,
    db:           Annotated[Session, Depends(get_db)],
):
    from app.models import AuditLog
    logs = (
        db.query(AuditLog)
        .filter(AuditLog.entity_type == "Expense", AuditLog.entity_id == expense_id)
        .order_by(AuditLog.created_at)
        .all()
    )
    return [
        {
            "id": l.id, "action": l.action, "changed_by": l.changed_by,
            "ip_address": l.ip_address, "created_at": l.created_at.isoformat(),
            "before": l.before_state, "after": l.after_state,
        }
        for l in logs
    ]


# ─── Internal helper ─────────────────────────────────────────────────────────

def _advance_chain(expense: Expense):
    """
    Move to the next step in the approval chain.
    If all steps are exhausted → APPROVED; otherwise stay VERIFIED and
    point current_approver_role at the next role.
    """
    chain = expense.approval_chain or []
    expense.approval_step += 1

    if expense.approval_step >= len(chain):
        # All approvers done
        expense.status                = ExpenseStatus.APPROVED
        expense.current_approver_role = None
    else:
        next_role = chain[expense.approval_step]
        # COST_CONTROL step already advances to VERIFIED
        if chain[expense.approval_step - 1] == RoleName.COST_CONTROL.value:
            expense.status = ExpenseStatus.VERIFIED
        expense.current_approver_role = next_role
