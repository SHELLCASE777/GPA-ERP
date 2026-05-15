"""
Import real GPA data:
  - PT Kertas Nusantara projects, values, and original PO/contract files
  - Inventory: consumables (LAPORAN STOCK CONSUMABLE GUDANG APRIL 2026.xlsx)
  - Inventory: tools (LAPORAN STOCK TOOLS GPA 2026.xlsx)

Run: python -m scripts.import_real_data
"""
import re
import sys
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
from pypdf import PdfReader
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database import SessionLocal
from app.models import InventoryItem, InventoryTxn, Project, ProjectDocument, ProjectStatus, User

CONSUMABLE_XLSX = Path(r"C:\Garuda\LAPORAN STOCK CONSUMABLE GUDANG APRIL 2026.xlsx")
TOOLS_XLSX = Path(r"C:\Garuda\LAPORAN STOCK TOOLS GPA 2026.xlsx")
KN_PROJECT_ROOT = Path(r"C:\Garuda\GPA\Commercial\PROJECT\WIN\PT KERTAS NUSANTARA\extracted")

PROJECTS = [
    ("GPA-KN-REVAMP", "Revamping PT Kertas Nusantara", "1.", "2025-01-01", "active"),
    ("P01.0825.316", "Additional Probe Shinkawa - PT KN", "2.", "2025-08-01", "active"),
    ("P01.0825.J075", "Repair & Services Control Valve & ON OFF Valve - PT KN", "3.", "2025-08-01", "active"),
    ("P01.0925.S489", "Replacement Instrument Defective MA 77 - PT KN", "4.", "2025-09-01", "active"),
    ("P01.1025.S549", "Deviation List Of Generator Overhaul 2x63MW - PT KN", "5.", "2025-10-01", "active"),
    ("P01.1025.S528", "Deviation Parts Turbine - PT KN", "6.", "2025-10-01", "active"),
    ("P01.1125.ELC", "Electrical Construction & Installation - PT KN", "7.", "2025-11-01", "active"),
    ("P01.1125.S680", "Overhaul Cooling Tower Pump Area MA 42 - PT KN", "8.", "2025-11-01", "active"),
    ("P01.1125.S690", "Proximity Switch Cable Scun & Ferrule MA 42/73 - PT KN", "9.", "2025-11-01", "active"),
    ("P01.1125.S689", "Cable Scun & Ferrule Terminal MA 42/73 - PT KN", "10.", "2025-11-01", "active"),
    ("P01.1125.S711", "Service & Repair Air Port Rodding Actuator MA 81 - PT KN", "11.", "2025-11-01", "active"),
    ("P01.1225.S788", "Replacement Temperature Sensor & Thermowell - PT KN", "12.", "2025-12-01", "active"),
    ("P01.1225.S793", "Toxic Gas Instrumentation - PT KN", "13.", "2025-12-01", "active"),
]


def map_category(raw: str | None) -> str:
    if not raw:
        return "consumables"
    r = str(raw).strip().upper()
    if r in {"TOOLS", "TOOL"}:
        return "tools"
    if r in {"MATERIAL", "MATERIALS"}:
        return "materials"
    return "consumables"


def clean_unit(unit: str | None) -> str:
    if not unit:
        return "pcs"
    return str(unit).strip().lower()


def safe_qty(value) -> Decimal:
    """The stock sheets store 10000 as 10 and 1000 as 1."""
    try:
        text = str(value).strip()
        if re.fullmatch(r"\d{1,3}(?:\.\d{3})+", text):
            text = text.replace(".", "")
        else:
            text = text.replace(",", ".")
        raw = float(text)
        qty = raw / 1000 if raw >= 1000 else raw
        return Decimal(str(max(0, qty)))
    except Exception:
        return Decimal("0")


def parse_id_amount(raw: str) -> Decimal | None:
    text = raw.strip().replace("Rp", "").replace("IDR", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "").replace(".", "")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except Exception:
        return None


def is_pdf(path: Path) -> bool:
    if path.suffix.lower() == ".pdf":
        return True
    try:
        with path.open("rb") as fh:
            return fh.read(5) == b"%PDF-"
    except Exception:
        return False


def is_project_doc(path: Path) -> bool:
    return path.suffix.lower() in {".pdf", ".xlsx", ".xls", ".docx"} or is_pdf(path)


def project_folder(prefix: str) -> Path | None:
    if not KN_PROJECT_ROOT.exists():
        return None
    return next((p for p in KN_PROJECT_ROOT.iterdir() if p.is_dir() and p.name.startswith(prefix)), None)


def extract_contract_value(folder: Path | None) -> Decimal:
    if not folder:
        return Decimal("0.00")
    amounts: list[Decimal] = []
    for path in folder.rglob("*"):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if is_pdf(path):
            try:
                pages = PdfReader(str(path)).pages
                text = "\n".join((page.extract_text() or "") for page in pages[:25])
            except Exception:
                continue
            for match in re.findall(r"(?:Rp\.?|IDR)?\s*([0-9]{1,3}(?:[.,][0-9]{3})+(?:[.,][0-9]{2})?)", text):
                amount = parse_id_amount(match)
                if amount and amount > 1_000_000:
                    amounts.append(amount)
        elif suffix in {".xlsx", ".xlsm"}:
            try:
                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            except Exception:
                continue
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for value in row:
                        if isinstance(value, (int, float)) and value > 1_000_000:
                            amounts.append(Decimal(str(value)).quantize(Decimal("0.01")))
                        elif isinstance(value, str):
                            for match in re.findall(r"(?:Rp\.?|IDR)\s*([0-9]{1,3}(?:[.,][0-9]{3})+(?:[.,][0-9]{2})?)", value, flags=re.I):
                                amount = parse_id_amount(match)
                                if amount and amount > 1_000_000:
                                    amounts.append(amount)
    return max(amounts, default=Decimal("0.00"))


def sync_project_documents(db: Session, project: Project, folder: Path | None) -> None:
    if not folder:
        return
    for path in folder.rglob("*"):
        if not path.is_file() or not is_project_doc(path):
            continue
        name_low = path.name.lower()
        doc_type = "po" if "p01." in name_low or " po" in name_low else (
            "contract" if "contract" in name_low or "perjanjian" in name_low else "proposal"
        )
        exists = db.query(ProjectDocument).filter(
            ProjectDocument.project_id == project.id,
            ProjectDocument.file_path == str(path),
        ).first()
        if exists:
            continue
        db.add(ProjectDocument(
            project_id=project.id,
            doc_type=doc_type,
            title=path.stem[:255],
            file_path=str(path),
            reference_no=project.code,
        ))


def read_consumables() -> list[dict]:
    wb = openpyxl.load_workbook(CONSUMABLE_XLSX, data_only=True)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=5, max_row=200, values_only=True):
        no, name, cat, brand, stock_in, unit, total_out = row[0], row[1], row[2], row[3], row[4], row[5], row[36]
        if not no or not name or name == 0:
            continue
        try:
            no = int(float(str(no)))
        except Exception:
            continue
        stock_in = safe_qty(stock_in)
        total_out = safe_qty(total_out)
        qty = max(Decimal("0"), stock_in - total_out)
        items.append({
            "code": f"CSM-{no:03d}",
            "name": str(name).strip().title(),
            "category": map_category(cat),
            "unit": clean_unit(unit),
            "qty_on_hand": qty,
            "stock_in": stock_in,
            "notes": f"Brand: {brand}" if brand and str(brand).strip() not in {"-", ""} else None,
        })
    return items


def read_tools() -> list[dict]:
    wb = openpyxl.load_workbook(TOOLS_XLSX, data_only=True)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=5, max_row=78, values_only=True):
        no, name, brand, stock, unit = row[0], row[1], row[2], row[3], row[4]
        if not no or not name or name == 0:
            continue
        try:
            no = int(float(str(no)))
        except Exception:
            continue
        qty = safe_qty(stock)
        items.append({
            "code": f"TLS-{no:03d}",
            "name": str(name).strip().title(),
            "category": "tools",
            "unit": clean_unit(unit),
            "qty_on_hand": qty,
            "stock_in": qty,
            "notes": f"Brand: {brand}" if brand and str(brand).strip() not in {"-", "0", ""} else None,
        })
    return items


def run():
    db: Session = SessionLocal()
    try:
        admin = db.query(User).filter(User.email == "admin@gpa.local").first() or db.query(User).first()
        if not admin:
            print("ERROR: No users found. Run seed.py first.")
            return

        print("\n=== PROJECTS ===")
        proj_created = proj_skipped = 0
        for code, name, folder_prefix, start, status_str in PROJECTS:
            folder = project_folder(folder_prefix)
            contract_value = extract_contract_value(folder)
            project = db.query(Project).filter(Project.code == code).first()
            if project:
                if project.contract_value == 0 and contract_value > 0:
                    project.contract_value = contract_value
                    project.currency = "IDR"
                sync_project_documents(db, project, folder)
                proj_skipped += 1
                print(f"  [=] {code} refreshed ({contract_value:,.2f})")
                continue
            project = Project(
                code=code,
                name=name,
                contract_value=contract_value,
                currency="IDR",
                status=ProjectStatus(status_str),
                start_date=date.fromisoformat(start),
            )
            db.add(project)
            db.flush()
            sync_project_documents(db, project, folder)
            proj_created += 1
            print(f"  [+] {code} - {name} ({contract_value:,.2f})")
        db.flush()
        print(f"  >> {proj_created} created, {proj_skipped} refreshed/skipped")

        print("\n=== INVENTORY ===")
        all_items = read_consumables() + read_tools()
        inv_created = inv_skipped = 0
        for data in all_items:
            existing = db.query(InventoryItem).filter(InventoryItem.code == data["code"]).first()
            if existing:
                existing.name = data["name"]
                existing.category = data["category"]
                existing.unit = data["unit"]
                existing.qty_on_hand = data["qty_on_hand"]
                existing.notes = data["notes"]
                existing.is_active = True
                inv_skipped += 1
                continue
            item = InventoryItem(
                code=data["code"],
                name=data["name"],
                category=data["category"],
                unit=data["unit"],
                qty_on_hand=data["qty_on_hand"],
                min_stock=Decimal("0"),
                notes=data["notes"],
                is_active=True,
            )
            db.add(item)
            db.flush()
            if data["stock_in"] > 0:
                db.add(InventoryTxn(
                    item_id=item.id,
                    txn_type="in",
                    quantity=data["stock_in"],
                    reference="OPENING",
                    notes="Saldo awal dari laporan April 2026",
                    created_by=admin.id,
                ))
            inv_created += 1
            print(f"  [+] {data['code']} {data['name']} - stok: {data['qty_on_hand']} {data['unit']}")

        db.commit()
        print(f"\nDone. {inv_created} inventory items created, {inv_skipped} skipped")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    run()
