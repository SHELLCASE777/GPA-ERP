"""
GPA-ERP V5.0 — Reporting Templates router.

Endpoints:
    GET /reports/payroll-summary      → Excel payroll summary for a given year/month
    GET /reports/project-financial    → Excel project financial report
"""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Annotated

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.dependencies import CurrentUser, require_role
from app.models import (
    Employee, Expense, ExpenseStatus, PayrollPeriod, PayrollRun,
    Project, ProjectStatus, RoleName,
)

router = APIRouter(prefix="/reports", tags=["Reports"])

# ── Styling helpers ────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="1E293B")
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _apply_headers(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill  = _HEADER_FILL
        cell.font  = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _autosize(ws, headers: list[str]) -> None:
    col_widths = [len(h) for h in headers]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                col_widths[cell.column - 1] = max(col_widths[cell.column - 1], len(str(cell.value)))
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width + 4


def _streaming(buf: io.BytesIO, filename: str) -> StreamingResponse:
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── GET /reports/payroll-summary ──────────────────────────────────────────────

_PAYROLL_ROLES = (RoleName.FINANCE, RoleName.MD, RoleName.SUPER_ADMIN)

MONTH_NAMES = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]


@router.get("/payroll-summary", summary="Export payroll summary to XLSX")
def export_payroll_summary(
    current_user: Annotated[object, Depends(require_role(*_PAYROLL_ROLES))],
    db:           Annotated[Session, Depends(get_db)],
    year:         int = Query(..., ge=2000, le=2100),
    month:        int = Query(..., ge=1,    le=12),
):
    period = (
        db.query(PayrollPeriod)
        .filter(PayrollPeriod.year == year, PayrollPeriod.month == month)
        .first()
    )
    if not period:
        raise HTTPException(status_code=404, detail=f"No payroll period found for {year}-{month:02d}")

    runs = (
        db.query(PayrollRun)
        .filter(PayrollRun.period_id == period.id)
        .options(
            joinedload(PayrollRun.employee).joinedload(Employee.department),
        )
        .order_by(PayrollRun.employee_id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Payroll"

    headers = [
        "No", "Bulan", "No Karyawan", "Nama", "Departemen",
        "Gaji Pokok", "Tunjangan", "OT Pay",
        "BPJS TK (Karyawan)", "BPJS Kes (Karyawan)", "PPh21",
        "Gaji Bersih",
    ]
    _apply_headers(ws, headers)

    bulan_label = f"{MONTH_NAMES[month]} {year}"

    for row_idx, run in enumerate(runs, start=1):
        emp  = run.employee
        snap = run.components_snapshot or {}

        # Pull component totals from snapshot
        basic_pay   = Decimal("0")
        allowance   = Decimal("0")
        ot_pay      = Decimal("0")

        for comp in snap.values() if isinstance(snap, dict) else []:
            comp_type = (comp.get("type") or "").upper()
            amount    = Decimal(str(comp.get("amount", 0)))
            if comp_type == "BASIC":
                basic_pay += amount
            elif comp_type == "ALLOWANCE":
                allowance += amount
            elif "OT" in (comp.get("code") or "").upper() or "OVERTIME" in (comp.get("name") or "").upper():
                ot_pay += amount

        dept_name = emp.department.name if emp and emp.department else ""

        ws.cell(row=row_idx + 1, column=1,  value=row_idx)
        ws.cell(row=row_idx + 1, column=2,  value=bulan_label)
        ws.cell(row=row_idx + 1, column=3,  value=emp.employee_no if emp else "")
        ws.cell(row=row_idx + 1, column=4,  value=emp.full_name   if emp else "")
        ws.cell(row=row_idx + 1, column=5,  value=dept_name)
        ws.cell(row=row_idx + 1, column=6,  value=float(basic_pay))
        ws.cell(row=row_idx + 1, column=7,  value=float(allowance))
        ws.cell(row=row_idx + 1, column=8,  value=float(ot_pay))
        ws.cell(row=row_idx + 1, column=9,  value=float(run.bpjs_tk_employee))
        ws.cell(row=row_idx + 1, column=10, value=float(run.bpjs_kes_employee))
        ws.cell(row=row_idx + 1, column=11, value=float(run.pph21_amount))
        ws.cell(row=row_idx + 1, column=12, value=float(run.net_salary))

    _autosize(ws, headers)

    buf = io.BytesIO()
    wb.save(buf)
    return _streaming(buf, f"payroll-summary-{year}-{month:02d}.xlsx")


# ── GET /reports/project-financial ────────────────────────────────────────────

_PROJECT_ROLES = (RoleName.MD, RoleName.COST_CONTROL, RoleName.FINANCE, RoleName.SUPER_ADMIN)

_PAID_STATUSES = [ExpenseStatus.PAID.value, ExpenseStatus.HARD_LOCKED.value]


@router.get("/project-financial", summary="Export project financial report to XLSX")
def export_project_financial(
    current_user: Annotated[object, Depends(require_role(*_PROJECT_ROLES))],
    db:           Annotated[Session, Depends(get_db)],
    status:       str | None = Query(None),
    year:         int | None = Query(None),
):
    q = db.query(Project).filter(Project.is_archived == False)

    if status:
        try:
            q = q.filter(Project.status == ProjectStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")

    if year:
        from sqlalchemy import extract
        q = q.filter(
            (Project.start_date == None) |
            (extract("year", Project.start_date) == year)
        )

    projects = q.order_by(Project.id).all()

    # Pre-fetch paid expense sums per project in one query
    paid_sums = dict(
        db.query(Expense.project_id, func.coalesce(func.sum(Expense.amount), 0))
        .filter(
            Expense.status.in_(_PAID_STATUSES),
            Expense.project_id.isnot(None),
        )
        .group_by(Expense.project_id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Keuangan Proyek"

    headers = [
        "No", "Kode Proyek", "Nama Proyek", "Status",
        "Nilai Kontrak", "Budget Terpakai", "% Burn Rate",
    ]
    _apply_headers(ws, headers)

    for row_idx, project in enumerate(projects, start=1):
        contract_value = float(project.contract_value or 0)
        budget_used    = float(paid_sums.get(project.id, Decimal("0")))
        burn_rate      = round(budget_used / contract_value * 100, 2) if contract_value else 0.0

        ws.cell(row=row_idx + 1, column=1, value=row_idx)
        ws.cell(row=row_idx + 1, column=2, value=project.code)
        ws.cell(row=row_idx + 1, column=3, value=project.name)
        ws.cell(row=row_idx + 1, column=4, value=project.status.value if project.status else "")
        ws.cell(row=row_idx + 1, column=5, value=contract_value)
        ws.cell(row=row_idx + 1, column=6, value=budget_used)
        ws.cell(row=row_idx + 1, column=7, value=burn_rate)

    _autosize(ws, headers)

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"project-financial-{year}.xlsx" if year else f"project-financial-{date.today().year}.xlsx"
    return _streaming(buf, filename)
